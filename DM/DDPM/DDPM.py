import numpy as np
import torch
import torch.nn as nn
import torch.optim as optim
import torchvision
import torchvision.models as models
import torchvision.transforms as TF
import matplotlib.pyplot as plt
import time
import torch.nn.functional as F
import copy
import random
from torch.utils.data import DataLoader, Subset, TensorDataset, ConcatDataset
from tqdm import tqdm
################################################
import gc
import os
import cv2
import math
import base64
from PIL import Image
from datetime import datetime
from torch.cuda import amp
from torch.optim import Adam, AdamW
import torchvision.datasets as datasets
from torchvision.utils import make_grid
from torchmetrics import MeanMetric

from IPython.display import display, HTML, clear_output

from torchvision.utils import save_image
#######################################################################
from torchvision.models import vgg16
from collections import namedtuple
import os
import requests
import shutil
from torchvision.utils import save_image
#######################################################################
from dataclasses import dataclass
import re
#######################################################################
import openpyxl
from openpyxl.styles import Font, Alignment
URL_MAP = {
    "vgg_lpips": "https://heibox.uni-heidelberg.de/f/607503859c864bc1b30b/?dl=1"
}

CKPT_MAP = {
    "vgg_lpips": "vgg.pth"
}
def download(url, local_path, chunk_size=1024):
    os.makedirs(os.path.split(local_path)[0], exist_ok=True)
    with requests.get(url, stream=True) as r:
        total_size = int(r.headers.get("content-length", 0))
        with tqdm(total=total_size, unit="B", unit_scale=True) as pbar:
            with open(local_path, "wb") as f:
                for data in r.iter_content(chunk_size=chunk_size):
                    if data:
                        f.write(data)
                        pbar.update(chunk_size)


device_idx = 3
os.environ["CUDA_VISIBLE_DEVICES"] = str(device_idx)
# 参数列表
K = 20 #用户数量
# select_nums = 5 #参与联邦训练的用户数量
# 设置随机数种子
np.random.seed(46) 
batch_size = 32
batch_size_test = 256
batch_size_pseudo = 256
batch_size_dm = 256
num_classes = 10

classify_num_rounds_lc = 200
classify_num_rounds_1 = 1000
classify_num_rounds_2 = 200
classify_num_rounds_fd = classify_num_rounds_1 + classify_num_rounds_2
classify_num_rounds_fd2 = 2000
classify_num_epochs_1 = 1
classify_num_epochs_2 = 5

dm_num_rounds_lc = 50
dm_num_rounds_1 = 400
dm_num_rounds_2 = 50
dm_num_rounds_fd = dm_num_rounds_1 + dm_num_rounds_2
dm_num_epochs_1 = 5
dm_num_epochs_2 = 20

vae_num_rounds_lc = 500
vae_num_rounds_fd = 400
vae_num_epochs = 3
latent_c = 6
downsample_factor = 1
latent_h = 32//(2*downsample_factor)
latent_w = 32//(2*downsample_factor)

discfactor = 1.
disc_start = 1
beta1= 0.5
beta2= 0.9
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")


classes = ('plane', 'car', 'bird', 'cat',
           'deer', 'dog', 'frog', 'horse', 'ship', 'truck')
## configurations


def get_default_device():
    return torch.device("cuda" if torch.cuda.is_available() else "cpu")
@dataclass
class BaseConfig:
    DEVICE = get_default_device()
    DATASET = "Cifar-10"  # "MNIST", "Cifar-10", "Cifar-100", "Flowers"

    # For logging inferece images and saving checkpoints.
    root_log_dir = os.path.join("Logs_Checkpoints", "Inference")
    root_checkpoint_dir = os.path.join("Logs_Checkpoints", "checkpoints")

    # Current log and checkpoint directory.
    log_dir = "version_0"
    checkpoint_dir = "version_0"


@dataclass
class TrainingConfig:
    TIMESTEPS = 1000  # Define number of diffusion timesteps
    # IMG_SHAPE = (1, 32, 32) if BaseConfig.DATASET == "MNIST" else (3, 32, 32)
    IMG_SHAPE = (latent_c, latent_h, latent_w)
    # NUM_EPOCHS = 800
    # BATCH_SIZE = 512
    LR = 2e-4 #2e-4
    # NUM_WORKERS = 0


def scale_to_range(t):
    # return (t * 2) - 1
    return (t + 1)*0.5

class VAE(nn.Module):
    def __init__(self, latent_c, downsample=1):
        super(VAE, self).__init__()

        self.latent_c = latent_c
        self.downsample = downsample

        self.perceptual_loss = LPIPS().eval().to(device=device)

        # 编码器
        self.encoder = nn.Sequential(
            nn.Conv2d(3, 32, kernel_size=3, stride=1, padding=1),
            nn.ReLU(),
            nn.MaxPool2d(kernel_size=2, stride=2)  # First downsampling by factor of 2
        )

        # 根据 downsample 增加额外的池化层
        for _ in range(self.downsample - 1):
            self.encoder.add_module("extra_pool_{}".format(_), nn.MaxPool2d(kernel_size=2, stride=2))

        # 均值和标准差卷积层
        self.fc_mu = nn.Conv2d(32, latent_c, kernel_size=1)
        self.fc_logvar = nn.Conv2d(32, latent_c, kernel_size=1)
        # # 解码器
        self.decoder = nn.Sequential()
        # 根据 downsample 添加相应数量的上采样+卷积层
        for _ in range(self.downsample - 1):
            # 添加双线性上采样模块
            self.decoder.add_module("upsample_{}".format(_),
                                    nn.Upsample(scale_factor=2, mode='bilinear', align_corners=True))
            # 添加卷积层以及ReLU激活层
            self.decoder.add_module("conv_{}".format(_),
                                    nn.Conv2d(latent_c, latent_c, kernel_size=3, stride=1, padding=1))
            self.decoder.add_module("relu_{}".format(_), nn.ReLU())

        # 最后的上采样和卷积层
        self.decoder.add_module("final_upsample", nn.Upsample(scale_factor=2, mode='bilinear', align_corners=True))
        self.decoder.add_module("final_conv", nn.Conv2d(latent_c, 32, kernel_size=3, stride=1, padding=1))
        self.decoder.add_module("final_relu", nn.ReLU())

        # 输出卷积层
        self.decoder.add_module("output_conv", nn.Conv2d(32, 3, kernel_size=3, stride=1, padding=1))
        self.decoder.add_module("output_sigmoid", nn.Sigmoid())

        # 判别器
        self.discriminator = Discriminator(image_channels=3)

    def reparameterize(self, mu, logvar):
        std = torch.exp(0.5 * logvar)
        eps = torch.randn_like(std)
        z = mu + eps * std
        return z

    def forward(self, x):
        # 通过编码器
        encoded = self.encoder(x)

        # 提取均值和对数方差
        mu = self.fc_mu(encoded)
        logvar = self.fc_logvar(encoded)

        # 采样潜在向量
        z = self.reparameterize(mu, logvar)

        # 通过解码器还原图像
        reconstructed = self.decoder(z)

        # 调整输出为输入的尺寸
        reconstructed = F.interpolate(reconstructed, size=(x.size(2), x.size(3)), mode='nearest')

        return reconstructed, mu, logvar

    def encode(self, x):
        encoded = self.encoder(x)
        mu = self.fc_mu(encoded)
        logvar = self.fc_logvar(encoded)
        z = self.reparameterize(mu, logvar)
        return z

    def decode(self, z):
        reconstructed = self.decoder(z)
        return reconstructed

    def discriminator_forward(self, x):
        # 通过判别器进行前向传播
        return self.discriminator(x)

    @staticmethod
    def adopt_weight(disc_factor, i, threshold, value=0.):
        if i < threshold:
            disc_factor = value
        return disc_factor

    def calculate_lambda(self, perceptual_loss, gan_loss):
        # 在这里，我们假设解码器的最后一个卷积层是输出层前的卷积层，即 'output_conv'
        last_layer = self.decoder.output_conv

        # 获取最后一层的权重
        last_layer_weight = last_layer.weight

        # 计算感知损失与GAN损失关于这个权重的梯度
        perceptual_loss_grads = torch.autograd.grad(perceptual_loss, last_layer_weight, retain_graph=True)[0]
        gan_loss_grads = torch.autograd.grad(gan_loss, last_layer_weight, retain_graph=True)[0]

        # 计算λ（lambda）的值
        λ = torch.norm(perceptual_loss_grads) / (torch.norm(gan_loss_grads) + 1e-4)  # 添加小数避免除零
        λ = torch.clamp(λ, 0, 1e4).detach()  # 限制λ的范围并从计算图中分离

        # 返回调整后的λ值
        return 0.8 * λ


class Discriminator(nn.Module):
    def __init__(self, image_channels, num_filters_last=64, n_layers=3):
        super(Discriminator, self).__init__()

        layers = [nn.Conv2d(image_channels, num_filters_last, 3, 2, 1), nn.LeakyReLU(0.2)]
        num_filters_mult = 1

        for i in range(1, n_layers + 1):
            num_filters_mult_last = num_filters_mult
            num_filters_mult = min(2 ** i, 8)
            layers += [
                nn.Conv2d(num_filters_last * num_filters_mult_last, num_filters_last * num_filters_mult, 4,
                          2 if i < n_layers else 1, 1, bias=False),
                nn.BatchNorm2d(num_filters_last * num_filters_mult),
                nn.LeakyReLU(0.2, True)
            ]

        layers.append(nn.Conv2d(num_filters_last * num_filters_mult, 1, 4, 1, 1))
        self.model = nn.Sequential(*layers)

    def forward(self, x):
        return self.model(x)

class LPIPS(nn.Module):
    def __init__(self):
        super(LPIPS, self).__init__()
        self.scaling_layer = ScalingLayer()
        self.channels = [64, 128, 256, 512, 512]
        self.vgg = VGG16()
        self.lins = nn.ModuleList([
            NetLinLayer(self.channels[0]),
            NetLinLayer(self.channels[1]),
            NetLinLayer(self.channels[2]),
            NetLinLayer(self.channels[3]),
            NetLinLayer(self.channels[4])
        ])

        self.load_from_pretrained()

        for param in self.parameters():
            param.requires_grad = False

    def load_from_pretrained(self, name="vgg_lpips"):
        ckpt = get_ckpt_path(name, "vgg_lpips")
        self.load_state_dict(torch.load(ckpt, map_location=torch.device("cpu")), strict=False)

    def forward(self, real_x, fake_x):
        features_real = self.vgg(self.scaling_layer(real_x))
        features_fake = self.vgg(self.scaling_layer(fake_x))
        diffs = {}

        for i in range(len(self.channels)):
            diffs[i] = (norm_tensor(features_real[i]) - norm_tensor(features_fake[i])) ** 2

        s = sum([spatial_average(self.lins[i].model(diffs[i])) for i in range(len(self.channels))])


        return s

class ScalingLayer(nn.Module):
    def __init__(self):
        super(ScalingLayer, self).__init__()
        self.register_buffer("shift", torch.Tensor([-.030, -.088, -.188])[None, :, None, None])
        self.register_buffer("scale", torch.Tensor([.458, .448, .450])[None, :, None, None])

    def forward(self, x):
        return (x - self.shift) / self.scale


class NetLinLayer(nn.Module):
    def __init__(self, in_channels, out_channels=1):
        super(NetLinLayer, self).__init__()
        self.model = nn.Sequential(
            nn.Dropout(),
            nn.Conv2d(in_channels, out_channels, 1, 1, 0, bias=False)
        )



class VGG16(nn.Module):
    def __init__(self):
        super(VGG16, self).__init__()
        vgg_pretrained_features = vgg16(pretrained=True).features
        slices = [vgg_pretrained_features[i] for i in range(30)]
        self.slice1 = nn.Sequential(*slices[0:4])
        self.slice2 = nn.Sequential(*slices[4:9])
        self.slice3 = nn.Sequential(*slices[9:16])
        self.slice4 = nn.Sequential(*slices[16:23])
        self.slice5 = nn.Sequential(*slices[23:30])

        for param in self.parameters():
            param.requires_grad = False

    def forward(self, x):
        h = self.slice1(x)
        h_relu1 = h
        h = self.slice2(h)
        h_relu2 = h
        h = self.slice3(h)
        h_relu3 = h
        h = self.slice4(h)
        h_relu4 = h
        h = self.slice5(h)
        h_relu5 = h
        vgg_outputs = namedtuple("VGGOutputs", ['relu1_2', 'relu2_2', 'relu3_3', 'relu4_3', 'relu5_3'])
        return vgg_outputs(h_relu1, h_relu2, h_relu3, h_relu4, h_relu5)


def get_ckpt_path(name, root):
    assert name in URL_MAP
    path = os.path.join(root, CKPT_MAP[name])
    if not os.path.exists(path):
        print(f"Downloading {name} model from {URL_MAP[name]} to {path}")
        download(URL_MAP[name], path)
    return path

def norm_tensor(x):
    """
    Normalize images by their length to make them unit vector?
    :param x: batch of images
    :return: normalized batch of images
    """
    norm_factor = torch.sqrt(torch.sum(x**2, dim=1, keepdim=True))
    return x / (norm_factor + 1e-10)

def spatial_average(x):
    """
     imgs have: batch_size x channels x width x height --> average over width and height channel
    :param x: batch of images
    :return: averaged images along width and height
    """
    return x.mean([2, 3], keepdim=True)


def train_vae(model, user_id, user_data, num_epochs, vae_optimizer, disc_optimizer, save_path):
    model.to(device)
    criterion = nn.BCELoss(reduction='mean')
    data_loader = DataLoader(Subset(trainset_vae, user_data), batch_size=128, shuffle=True)
    steps_per_epoch = len(data_loader)

    # 创建每个用户的结果和检查点目录
    user_results_dir = os.path.join(save_path,f"Local/vae_results/user_{user_id+1}" )
    user_checkpoints_dir = os.path.join(save_path, f"Local/vae_checkpoints/user_{user_id+1}")
    os.makedirs(user_results_dir, exist_ok=True)
    os.makedirs(user_checkpoints_dir, exist_ok=True)

    # 目标目录
    dest_dir = os.path.join(save_path, f"Local/vae_model")
    os.makedirs(dest_dir, exist_ok=True)  # 确保目标目录存在

    model.train()

    for epoch in range(num_epochs):
        pbar = tqdm(enumerate(data_loader), total=steps_per_epoch, desc=f"Epoch {epoch + 1}/{num_epochs}")
        for batch_idx, (data, _) in pbar:
            data = data.to(device, dtype=torch.float)

            recon_batch, mu, logvar = model(data)
            disc_real = model.discriminator(data)
            disc_fake = model.discriminator(recon_batch)
            disc_factor = model.adopt_weight(discfactor, epoch * steps_per_epoch + batch_idx, threshold=disc_start)
            loss_perceptual = model.perceptual_loss(data, recon_batch)  # 感知损失

            loss_recon = torch.abs(data - recon_batch)  # 重构损失
            loss_per_recon = loss_perceptual + loss_recon
            loss_per_recon = loss_per_recon.mean()
            # loss_recon = criterion(recon_batch, data)

            loss_kl = -0.5 * torch.sum(1 + logvar - mu.pow(2) - logvar.exp())
            perceptual_rec_kl_loss = loss_per_recon
            # perceptual_rec_kl_loss = perceptual_rec_kl_loss.mean()
            g_loss = -torch.mean(disc_fake)

            λ = model.calculate_lambda(perceptual_rec_kl_loss, g_loss)
            vae_loss = perceptual_rec_kl_loss + discfactor * λ * g_loss
            d_loss_real = torch.mean(F.relu(1. - disc_real))
            d_loss_fake = torch.mean(F.relu(1. + disc_fake))
            gan_loss = disc_factor * 0.5 * (d_loss_real + d_loss_fake)

            vae_optimizer.zero_grad()
            vae_loss.backward(retain_graph=True)
            disc_optimizer.zero_grad()
            gan_loss.backward()
            vae_optimizer.step()
            disc_optimizer.step()
            # vae_optimizer.zero_grad()
            # vae_loss.backward(retain_graph=True)
            # vae_optimizer.step()
            #
            # disc_optimizer.zero_grad()
            # gan_loss.backward()
            # disc_optimizer.step()

            # Update progress bar
            pbar.set_postfix(VAE_Loss=vae_loss.item(), GAN_Loss=gan_loss.item())

            # Save images conditionally
            if batch_idx % 100 == 0:
                with torch.no_grad():
                    comparison = torch.cat((data[:4], recon_batch[:4]))
                    save_image(comparison, os.path.join(user_results_dir, f"recon_epoch{epoch+1}_batch{batch_idx}.png"), nrow=4)

        # Save model checkpoints for each epoch
        if(epoch % 10 == 0):
            model_checkpoint_path = os.path.join(user_checkpoints_dir, f"vae_epoch_{epoch+1}.pt")
            torch.save(model.state_dict(), model_checkpoint_path)

    # 复制最后一个模型到新的文件夹并重命名
    final_model_path = os.path.join(dest_dir, f"final_vae_model_{user_id + 1}.pt")
    shutil.copy2(model_checkpoint_path, final_model_path)
    print(f"Copied final model for user {user_id + 1} to {final_model_path}")

def train_vae_fed(model, user_id, user_data, num_epochs, vae_optimizer, disc_optimizer, current_round, save_path):
    model.to(device)
    criterion = nn.BCELoss(reduction='mean')
    data_loader = DataLoader(Subset(trainset_vae, user_data), batch_size=256, shuffle=True)
    steps_per_epoch = len(data_loader)

    # 创建每个用户的结果目录
    user_results_dir = os.path.join(save_path, f"Fed/vae_results/user_{user_id+1}")
    os.makedirs(user_results_dir, exist_ok=True)


    model.train()

    for epoch in range(num_epochs):
        pbar = tqdm(enumerate(data_loader), total=steps_per_epoch, desc=f"Epoch {epoch + 1}/{num_epochs}, Round:{current_round + 1 }")
        for batch_idx, (data, _) in pbar:
            data = data.to(device, dtype=torch.float)

            recon_batch, mu, logvar = model(data)
            disc_real = model.discriminator(data)
            disc_fake = model.discriminator(recon_batch)
            disc_factor = model.adopt_weight(discfactor, epoch * steps_per_epoch + batch_idx, threshold=disc_start)
            loss_perceptual = model.perceptual_loss(data, recon_batch)  # 感知损失

            loss_recon = torch.abs(data - recon_batch)  # 重构损失
            loss_per_recon = loss_perceptual + loss_recon
            loss_per_recon = loss_per_recon.mean()
            # loss_recon = criterion(recon_batch, data)

            loss_kl = -0.5 * torch.sum(1 + logvar - mu.pow(2) - logvar.exp())
            perceptual_rec_kl_loss = loss_per_recon
            # perceptual_rec_kl_loss = perceptual_rec_kl_loss.mean()
            g_loss = -torch.mean(disc_fake)

            λ = model.calculate_lambda(perceptual_rec_kl_loss, g_loss)
            vae_loss = perceptual_rec_kl_loss + discfactor * λ * g_loss
            d_loss_real = torch.mean(F.relu(1. - disc_real))
            d_loss_fake = torch.mean(F.relu(1. + disc_fake))
            gan_loss = disc_factor * 0.5 * (d_loss_real + d_loss_fake)

            vae_optimizer.zero_grad()
            vae_loss.backward(retain_graph=True)
            disc_optimizer.zero_grad()
            gan_loss.backward()
            vae_optimizer.step()
            disc_optimizer.step()
            # vae_optimizer.zero_grad()
            # vae_loss.backward(retain_graph=True)
            # vae_optimizer.step()
            #
            # disc_optimizer.zero_grad()
            # gan_loss.backward()
            # disc_optimizer.step()

            # Update progress bar
            pbar.set_postfix(VAE_Loss=vae_loss.item(), GAN_Loss=gan_loss.item())

            # Save images conditionally
            if (current_round + 1) % 25 == 0 and epoch == num_epochs -1 and batch_idx == len(data_loader) - 1: 
                with torch.no_grad():
                    comparison = torch.cat((data[:4], recon_batch[:4]))
                    save_image(comparison, os.path.join(user_results_dir,f"recon_epoch{epoch + 1}_batch{batch_idx + 1}_round{current_round + 1}.png"),nrow=4)

        # Save model checkpoints for each epoch
        # if(current_round == 0) or ((current_round+1) % 50 == 0):
        #     if(epoch == 0) or ((epoch+1) % 10 == 0):
        #         user_checkpoints_dir = os.path.join(save_path, f"Fed/vae_checkpoints/user_{user_id + 1}/round_{current_round + 1}")
        #         # user_checkpoints_dir = f"FD-VAE-GAN/Fed/vae_checkpoints/user_{user_id + 1}/round_{current_round + 1}"
        #         os.makedirs(user_checkpoints_dir, exist_ok=True)
        #         model_checkpoint_path = os.path.join(user_checkpoints_dir, f"vae_epoch_{epoch + 1}_round{current_round + 1 }.pt")
        #         torch.save(model.state_dict(), model_checkpoint_path)

    # 记录聚合后过程中的模型，复制最后一个epoch的模型到新的文件夹并重命名
    # if (current_round == 0) or ((current_round + 1) % 50 == 0):
    #     # 目标目录
    #     dest_dir = os.path.join(save_path, f"Fed/vae_model/user_{user_id + 1}")
    #     # dest_dir = f"FD-VAE-GAN/Fed/vae_model/user_{user_id + 1}"
    #     os.makedirs(dest_dir, exist_ok=True)  # 确保目标目录存在
    #     final_model_path = os.path.join(dest_dir, f"final_vae_model_{user_id + 1}_round_{current_round + 1}.pt")
    #     shutil.copy2(model_checkpoint_path, final_model_path)
    #     print(f"Copied round_{current_round+1} model for user {user_id + 1} to {final_model_path}")
    if (current_round == 0) or ((current_round + 1) % 100 == 0):
        dest_dir = os.path.join(save_path, f"Fed/vae_model/user_{user_id + 1}")
        os.makedirs(dest_dir, exist_ok=True)
        final_model_path = os.path.join(dest_dir, f"final_vae_model_{user_id + 1}_round_{current_round + 1}.pt")
        torch.save(model.state_dict(), final_model_path)
        print(f"Saved final model for user {user_id + 1} round {current_round + 1} to {final_model_path}")
    
    return model.state_dict()

def federated_average(models, data_sizes):
    global_model = copy.deepcopy(models[0])
    total_data_size = sum(data_sizes)
    # Initialize global model parameters to zero
    for key in global_model.keys():
        global_model[key] = torch.zeros_like(global_model[key]).float()

    for model, data_size in zip(models, data_sizes):
        weight = data_size / total_data_size
        for key in global_model.keys():
            # Ensure the parameters are in float for precise calculations
            global_model[key] += model[key].float() * weight
    return global_model

def create_incremental_folder(base_path):
    """
    创建一个数字递增命名的新子文件夹。

    Args:
    base_path (str): 基础路径，函数将在此路径下创建新的数字命名文件夹。

    Returns:
    str: 新创建的子文件夹的完整路径。
    """
    # 确保base_path指向的目录存在
    os.makedirs(base_path, exist_ok=True)

    # 获取所有数字命名的子文件夹并找出最大的数字
    max_num = 0
    with os.scandir(base_path) as entries:
        for entry in entries:
            if entry.is_dir() and entry.name.isdigit():
                num = int(entry.name)
                if num > max_num:
                    max_num = num

    # 根据当前最大数字决定新文件夹的名称
    new_folder_name = str(max_num + 1)

    # 创建新文件夹
    new_folder_path = os.path.join(base_path, new_folder_name)
    os.makedirs(new_folder_path, exist_ok=True)

    # 返回新创建的文件夹路径
    return new_folder_path
# vae_model_list=[]
# vae_optimizer_list=[]
# for _ in range(K):
#     vae_model = VAE(latent_h, latent_c, latent_w).to(device)
#     vae_optimizer = torch.optim.Adam(vae_model.parameters(), lr=1e-3)
#     vae_model_list.append(vae_model)
#     vae_optimizer_list.append(vae_optimizer)


# # 训练每个用户的VAE
# for user_id, user_data in enumerate(train_user_datasets_labelled):
#     print(f"Training VAE for User {user_id + 1}")
#     labelled_data = user_data
#     unlabelled_data = train_user_datasets_unlabelled[user_id]
#     combined_data = labelled_data + unlabelled_data
#     train_vae(vae_model_list[user_id], combined_data, vae_num_rounds_lc, vae_optimizer_list[user_id])
#
# ###训好的每个用户的vae保存在vae_model_list[user_id]


def show_images(original_images, reconstructed_images, num_images=10):
    fig, axes = plt.subplots(num_images, 2, figsize=(8, 2 * num_images))
    for i in range(num_images):
        # 显示原始图像
        axes[i, 0].imshow(np.transpose(original_images[i], (1, 2, 0)))
        axes[i, 0].set_title('Original Image')
        axes[i, 0].axis('off')

        # 显示重构图像
        axes[i, 1].imshow(np.transpose(reconstructed_images[i], (1, 2, 0)))
        axes[i, 1].set_title('Reconstructed Image')
        axes[i, 1].axis('off')

    plt.tight_layout()
    plt.show()

# # 测试VAE模型
# for user_id, user_data in enumerate(test_user_datasets_vae):
#     vae_test_model = vae_model_list[user_id]
#     vae_test_loader = DataLoader(Subset(testset_vae, user_data), batch_size=64, shuffle=True)
#     vae_test_model.eval()
#     with torch.no_grad():
#         original_images = []
#         reconstructed_images = []
#         for batch_idx, (data, _) in enumerate(vae_test_loader):
#             data = data.to(device, dtype=torch.float)  # 将输入数据转换为与模型权重相同的类型和设备
#             recon_batch, _, _ = vae_test_model(data)
#             original_images.append(data.cpu().numpy())
#             reconstructed_images.append(recon_batch.cpu().numpy())
#
#         original_images = np.concatenate(original_images, axis=0)
#         reconstructed_images = np.concatenate(reconstructed_images, axis=0)
#
#     # 显示前10个原始图像和重构图像
#     show_images(original_images, reconstructed_images, num_images=10)
def load_model(model_path, model_architecture):
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    model = model_architecture.to(device)
    model.load_state_dict(torch.load(model_path))
    return model


def create_encoded_dataloader_with_labels(model, dataloader, device):
    model.eval()  # 确保模型处于评估模式
    encoded_features = []  # 存储所有编码向量
    labels_list = []  # 存储所有标签

    with torch.no_grad():  # 确保在推理过程中不计算梯度
        for images, labels in dataloader:
            images = images.to(device)  # 将图像数据移至正确的设备
            encoded = model.encode(images)  # 使用VAE的encode方法获取潜在向量
            # # 归一化到 -1 到 1
            # min_val = encoded.min()
            # max_val = encoded.max()
            # encoded = 2 * ((encoded - min_val) / (max_val - min_val)) - 1

            encoded_features.append(encoded)
            labels_list.append(labels)

    # 将所有编码向量和标签合并为一个tensor
    encoded_features = torch.cat(encoded_features, dim=0)
    labels = torch.cat(labels_list, dim=0)

    # 创建新的带标签的dataloader
    encoded_dataset = TensorDataset(encoded_features, labels)
    encoded_dataloader = DataLoader(encoded_dataset, batch_size=batch_size_dm, shuffle=False)

    return encoded_dataloader


def get_latest_model(directory, user_id):
    """
    返回指定目录下round数最大的模型文件路径。

    :param directory: 包含模型文件的目录路径
    :param user_id: 用户ID
    :return: round数最大的模型文件的完整路径，如果没有找到则返回None
    """
    # 构建文件名模式
    pattern = f"final_vae_model_{user_id}_round_(\d+)\.pt"

    latest_round = -1
    latest_model = None

    for filename in os.listdir(directory):
        match = re.match(pattern, filename)
        if match:
            round_num = int(match.group(1))
            if round_num > latest_round:
                latest_round = round_num
                latest_model = os.path.normpath(os.path.join(directory, filename))

    if latest_model:
        print(f"Latest model for user {user_id}: {latest_model} (Round {latest_round})")
    else:
        print(f"No models found for user {user_id}")

    return latest_model
#########################################################################################
####训练condition-DDPM
class SinusoidalPositionEmbeddings(nn.Module):
    def __init__(self, total_time_steps=1000, time_emb_dims=128, time_emb_dims_exp=512):
        super().__init__()

        half_dim = time_emb_dims // 2

        emb = math.log(10000) / (half_dim - 1)
        emb = torch.exp(torch.arange(half_dim, dtype=torch.float32) * -emb)

        ts = torch.arange(total_time_steps, dtype=torch.float32)

        emb = torch.unsqueeze(ts, dim=-1) * torch.unsqueeze(emb, dim=0)
        emb = torch.cat((emb.sin(), emb.cos()), dim=-1)

        self.time_blocks = nn.Sequential(
            nn.Embedding.from_pretrained(emb),
            nn.Linear(in_features=time_emb_dims, out_features=time_emb_dims_exp),
            nn.SiLU(),
            nn.Linear(in_features=time_emb_dims_exp, out_features=time_emb_dims_exp),
        )

    def forward(self, time):
        return self.time_blocks(time)

###标签嵌入类
class LabelEmbeddings(nn.Module):
    def __init__(self, num_classes, embed_dim):
        super().__init__()
        self.num_classes = num_classes
        self.embed_dim = embed_dim
        self.embedding = nn.Embedding(num_classes, embed_dim)

    def forward(self, labels):
        return self.embedding(labels)



class AttentionBlock(nn.Module):
    def __init__(self, channels=64, label_dim=128):
        super().__init__()
        self.channels = channels
        self.label_dim = label_dim

        self.group_norm_x = nn.GroupNorm(num_groups=8, num_channels=channels)  # 对输入 x 进行的组归一化
        self.group_norm_y = nn.GroupNorm(num_groups=8, num_channels=label_dim)  # 对输入 y 进行的组归一化
        self.label_proj = nn.Linear(label_dim, channels)  # 将 label_dim 投射到 channels 维度
        self.mhsa = nn.MultiheadAttention(embed_dim=channels, num_heads=4, batch_first=True)  # 多头注意力机制

    def forward(self, x, y=None):
        B, _, H, W = x.shape

        # 对 x 进行批归一化并调整形状
        h_x = self.group_norm_x(x)
        h_x = h_x.reshape(B, self.channels, H * W).swapaxes(1, 2)  # [B, C, H, W] -> [B, H * W, C]

        if y is None:
            # 自注意力
            h, _ = self.mhsa(h_x, h_x, h_x)  # Query, Key, Value 都是 h_x
        else:
            # 检查 y 的形状，如果是二维的，则调整形状
            if len(y.shape) == 2:
                y = y.unsqueeze(2).unsqueeze(3)  # [B, label_dim] -> [B, label_dim, 1, 1]
            elif len(y.shape) != 4:
                raise ValueError(f'Expected y to have 2 or 4 dimensions, but got {len(y.shape)}')

            # 对 y 进行批归一化并调整形状
            h_y = self.group_norm_y(y)
            h_y = self.label_proj(h_y.reshape(B, self.label_dim))  # [B, label_dim, 1, 1] -> [B, channels]
            h_y = h_y.unsqueeze(1)  # [B, channels] -> [B, 1, channels]

            # 交叉注意力
            h, _ = self.mhsa(h_x, h_y, h_y)  # Query 是 h_x，Key 和 Value 是 h_y

        # 将注意力输出 h 调整回原始形状
        h = h.swapaxes(1, 2).reshape(B, self.channels, H, W)  # [B, H * W, C] -> [B, C, H, W]

        # 返回 x + h，实现残差连接
        return x + h




class ResnetBlock(nn.Module):
    def __init__(self, *, in_channels, out_channels, dropout_rate=0.1, time_emb_dims=512,label_emb_dims=512, apply_attention=False):
        super().__init__()
        self.in_channels = in_channels
        self.out_channels = out_channels
        self.act_fn = nn.SiLU()  # 使用 SiLU 激活函数

        # Group 1
        self.normlize_1 = nn.GroupNorm(num_groups=8, num_channels=self.in_channels)  # 批归一化层
        self.conv_1 = nn.Conv2d(in_channels=self.in_channels, out_channels=self.out_channels,
                                kernel_size=3, stride=1, padding="same")  # 卷积层

        # Group 2 time embedding
        self.dense_1 = nn.Linear(in_features=time_emb_dims, out_features=self.out_channels)  # 全连接层

        # Group 3
        self.normlize_2 = nn.GroupNorm(num_groups=8, num_channels=self.out_channels)  # 批归一化层
        self.dropout = nn.Dropout2d(p=dropout_rate)  # 二维随机丢弃层
        self.conv_2 = nn.Conv2d(in_channels=self.out_channels, out_channels=self.out_channels,
                                kernel_size=3, stride=1, padding="same")  # 卷积层

        if self.in_channels != self.out_channels:
            self.match_input = nn.Conv2d(in_channels=self.in_channels, out_channels=self.out_channels,
                                         kernel_size=1, stride=1)  # 卷积层用于匹配输入和输出通道数
        else:
            self.match_input = nn.Identity()  # 恒等映射，用于通道数已匹配的情况

        if apply_attention:
            # self.attention = AttentionBlock(channels=self.out_channels)  # 注意力块
            self.attention = AttentionBlock(channels=self.out_channels,label_dim=label_emb_dims) # 交叉注意力块
        else:
            self.attention = nn.Identity()  # 不应用注意力机制，恒等映射

    def forward(self, x, t, y):
        ##y代表标签
        # group 1
        h = self.act_fn(self.normlize_1(x))
        h = self.conv_1(h)

        # group 2
        # add in timestep embedding
        h += self.dense_1(self.act_fn(t))[:, :, None, None]

        # group 3
        h = self.act_fn(self.normlize_2(h))
        h = self.dropout(h)
        h = self.conv_2(h)

        # Residual and attention
        h = h + self.match_input(x)
        if isinstance(self.attention, AttentionBlock):
            h = self.attention(h, y)
        else:
            h = self.attention(h)

        return h


class DownSample(nn.Module):
    def __init__(self, channels):
        super().__init__()
        self.downsample = nn.Conv2d(in_channels=channels, out_channels=channels, kernel_size=3, stride=2, padding=1)

    def forward(self, x, *args):
        return self.downsample(x)


class UpSample(nn.Module):
    def __init__(self, in_channels):
        super().__init__()

        self.upsample = nn.Sequential(
            nn.Upsample(scale_factor=2, mode="nearest"),
            nn.Conv2d(in_channels=in_channels, out_channels=in_channels, kernel_size=3, stride=1, padding=1),
        )

    def forward(self, x, *args):
        return self.upsample(x)


class UNet(nn.Module):
    def __init__(
        self,
        input_channels=3,  # 输入图像的通道数,默认为3(RGB)
        output_channels=3,  # 输出图像的通道数,默认为3(RGB)
        num_res_blocks=2,  # 每个级别的ResNet块的数量，Deocder层最终使用的UpBlock数=num_res_blocks + 1
        base_channels=128,  # 第一层卷积的输出通道数
        base_channels_multiples=(1, 2, 4, 8),  # 每个层的通道数相对于第一层卷积输出通道数的倍数
        apply_attention=(False, False, True, False),  # 在每个层是否应用注意力机制
        dropout_rate=0.1,  # ResNet块中使用的dropout率
        time_multiple=4,  # 时间嵌入维度相对于base_channels的倍数
        label_multiple=4,  # 标签嵌入维度相对于base_channels的倍数
        num_classes=10,


    ):
        super().__init__()

        # 计算时间嵌入的维度
        time_emb_dims_exp = base_channels * time_multiple
        # 创建正弦位置嵌入层
        self.time_embeddings = SinusoidalPositionEmbeddings(time_emb_dims=base_channels, time_emb_dims_exp=time_emb_dims_exp)

        # 计算标签嵌入的维度
        label_emb_dims_exp = base_channels * label_multiple
        ## 创建标签嵌入层
        self.label_embeddings = LabelEmbeddings(num_classes=num_classes, embed_dim=label_emb_dims_exp)

        # 第一个卷积层,将输入通道数转换为基础通道数
        self.first = nn.Conv2d(in_channels=input_channels, out_channels=base_channels, kernel_size=3, stride=1, padding="same")

        num_resolutions = len(base_channels_multiples)#分辨率数量，因为每次下采样或者上采样都会改变分辨率，也就是encoder和decoder的层数

        # UNet的编码器部分,进行维度缩减
        self.encoder_blocks = nn.ModuleList()  # 创建一个空的ModuleList,用于存储编码器部分的所有层
        curr_channels = [base_channels]  # 初始化一个列表,用于跟踪当前级别的通道数,初始值为base_channels,方便后续decoder确定通道数
        in_channels = base_channels  # 初始化当前级别的输入通道数为base_channels

        for level in range(num_resolutions):  # 对于每一层
            out_channels = base_channels * base_channels_multiples[level]  # 计算当前级别的输出通道数

            for _ in range(num_res_blocks):  # 对于每个ResNet块
                # 创建ResNet块
                block = ResnetBlock(
                    in_channels=in_channels,  # 当前ResNet块的输入通道数
                    out_channels=out_channels,  # 当前ResNet块的输出通道数
                    dropout_rate=dropout_rate,  # ResNet块中使用的dropout率
                    time_emb_dims=time_emb_dims_exp,  # 时间嵌入的维度
                    label_emb_dims=label_emb_dims_exp,  # 标签嵌入的维度
                    apply_attention=apply_attention[level],  # 是否在当前级别应用注意力机制
                )
                self.encoder_blocks.append(block)  # 将创建的ResNet块添加到编码器的ModuleList中

                in_channels = out_channels  # 更新下一个ResNet块的输入通道数为当前ResNet块的输出通道数
                curr_channels.append(in_channels)  # 将当前级别的输出通道数添加到curr_channels列表中,用于跟踪编码器每个级别的输出通道数

            if level != (num_resolutions - 1):
                # 在编码器的每个级别(除了最后一级)添加下采样层
                self.encoder_blocks.append(DownSample(channels=in_channels))
                curr_channels.append(in_channels)

        # 中间的瓶颈块
        self.bottleneck_blocks = nn.ModuleList(
            (
                ResnetBlock(
                    in_channels=in_channels,
                    out_channels=in_channels,
                    dropout_rate=dropout_rate,
                    time_emb_dims=time_emb_dims_exp,
                    label_emb_dims=label_emb_dims_exp,  # 标签嵌入的维度
                    apply_attention=True,
                ),
                ResnetBlock(
                    in_channels=in_channels,
                    out_channels=in_channels,
                    dropout_rate=dropout_rate,
                    time_emb_dims=time_emb_dims_exp,
                    label_emb_dims=label_emb_dims_exp,  # 标签嵌入的维度
                    apply_attention=False,
                ),
            )
        )

        # UNet的解码器部分,进行维度恢复,并使用跳跃连接
        self.decoder_blocks = nn.ModuleList()

        for level in reversed(range(num_resolutions)):
            out_channels = base_channels * base_channels_multiples[level]

            for _ in range(num_res_blocks + 1):#解码器每一层的resnet块数等于对应编码器层的块数+1
                encoder_in_channels = curr_channels.pop()
                # 创建ResNet块,并将编码器的输出与解码器的输入连接
                block = ResnetBlock(
                    in_channels=encoder_in_channels + in_channels,#跳跃连接
                    out_channels=out_channels,
                    dropout_rate=dropout_rate,
                    time_emb_dims=time_emb_dims_exp,
                    label_emb_dims=label_emb_dims_exp,  # 标签嵌入的维度
                    apply_attention=apply_attention[level],
                )

                in_channels = out_channels
                self.decoder_blocks.append(block)

            if level != 0:
                # 在解码器的每个级别(除了最后一级)添加上采样层
                self.decoder_blocks.append(UpSample(in_channels))

        # 最后的输出块,将通道数转换为输出通道数
        self.final = nn.Sequential(
            nn.GroupNorm(num_groups=8, num_channels=in_channels),
            nn.SiLU(),
            nn.Conv2d(in_channels=in_channels, out_channels=output_channels, kernel_size=3, stride=1, padding="same"),
        )

    def forward(self, x, y, t):
        # 获取时间步长的嵌入
        time_emb = self.time_embeddings(t)
        # 获取标签的嵌入
        label_emb = self.label_embeddings(y)


        # 第一个卷积层
        h = self.first(x)
        outs = [h]

        # 编码器部分
        for layer in self.encoder_blocks:
            if isinstance(layer, ResnetBlock):
                h = layer(h, time_emb, label_emb)
            else:
                h = layer(h, time_emb)
            outs.append(h)

        # 瓶颈块
        for layer in self.bottleneck_blocks:
            h = layer(h, time_emb, label_emb)

        # 解码器部分
        for layer in self.decoder_blocks:
            if isinstance(layer, ResnetBlock):#判断是否是resnetblock，跳跃连接，而上采用则不需要
                out = outs.pop()
                h = torch.cat([h, out], dim=1)  # 连接编码器的输出和解码器的输入
                h = layer(h, time_emb, label_emb)
            else:
                h = layer(h, time_emb)



        # 最后的输出块
        h = self.final(h)

        return h


####helper functions
def to_device(data, device):
    """Move tensor(s) to chosen device"""
    if isinstance(data, (list, tuple)):
        return [to_device(x, device) for x in data]
    return data.to(device, non_blocking=True)


class DeviceDataLoader:
    """Wrap a dataloader to move data to a device"""

    def __init__(self, dl, device):
        self.dl = dl
        self.device = device

    def __iter__(self):
        """Yield a batch of data after moving it to device"""
        for b in self.dl:
            yield to_device(b, self.device)

    def __len__(self):
        """Number of batches"""
        return len(self.dl)





def save_images(images, path, **kwargs):
    grid = make_grid(images, **kwargs)
    ndarr = grid.permute(1, 2, 0).to("cpu").numpy()
    im = Image.fromarray(ndarr)
    im.save(path)


def get(element: torch.Tensor, t: torch.Tensor):
    """
    Get value at index position "t" in "element" and
        reshape it to have the same dimension as a batch of images.
    """
    ele = element.gather(-1, t)
    return ele.reshape(-1, 1, 1, 1)





def frames2vid(images, save_path):
    WIDTH = images[0].shape[1]
    HEIGHT = images[0].shape[0]

    #     fourcc = cv2.VideoWriter_fourcc(*'XVID')
    #     fourcc = 0
    fourcc = cv2.VideoWriter_fourcc(*'mp4v')
    video = cv2.VideoWriter(save_path, fourcc, 25, (WIDTH, HEIGHT))

    # Appending the images to the video one by one
    for image in images:
        video.write(image)

    # Deallocating memories taken for window creation
    # cv2.destroyAllWindows()
    video.release()
    return


def display_gif(gif_path):
    b64 = base64.b64encode(open(gif_path, 'rb').read()).decode('ascii')
    display(HTML(f'<img src="data:image/gif;base64,{b64}" />'))



###数据集加载
# def scale_to_range(t):
#     return (t * 2) - 1


def get_dataset(dataset_name='MNIST'):
    transforms = TF.Compose(
        [
            TF.ToTensor(),
            TF.Resize((32, 32),
                      interpolation=TF.InterpolationMode.BICUBIC,
                      antialias=True),
            TF.RandomHorizontalFlip(),
            # TF.Lambda(lambda t: (t * 2) - 1) # Scale between [-1, 1]
            TF.Lambda(scale_to_range)
        ]
    )

    if dataset_name.upper() == "MNIST":
        dataset = datasets.MNIST(root="data", train=True, download=True, transform=transforms)
    elif dataset_name == "Cifar-10":
        dataset = datasets.CIFAR10(root="data", train=True, download=True, transform=transforms)
    elif dataset_name == "Cifar-100":
        dataset = datasets.CIFAR10(root="data", train=True, download=True, transform=transforms)
    elif dataset_name == "Flowers":
        dataset = datasets.ImageFolder(root=r"D:\Lib\wzw\myProject\FLDM\ddpm\flowers_dataset\dataset\train",
                                       transform=transforms)



    return dataset


def get_dataloader(dataset_name='MNIST',
                   batch_size=32,
                   pin_memory=False,
                   shuffle=True,
                   num_workers=0,
                   device="cpu"
                   ):
    dataset = get_dataset(dataset_name=dataset_name)
    dataloader = DataLoader(dataset, batch_size=batch_size,
                            pin_memory=pin_memory,
                            num_workers=num_workers,
                            shuffle=shuffle
                            )
    device_dataloader = DeviceDataLoader(dataloader, device)
    return device_dataloader


def inverse_transform(tensors):
    """Convert tensors from [-1., 1.] to [0., 255.]"""
    # return ((tensors.clamp(-1, 1) + 1.0) / 2.0) * 255.0
    return tensors * 255.0

# ###可视化数据集
# loader = get_dataloader(
#     dataset_name=BaseConfig.DATASET,
#     batch_size=128,
#     device='cpu',
# )
#
# plt.figure(figsize=(12, 6), facecolor='white')
#
# for b_image, _ in loader:
#     b_image = inverse_transform(b_image).cpu()
#     grid_img = make_grid(b_image / 255.0, nrow=16, padding=True, pad_value=1, normalize=True)
#     plt.imshow(grid_img.permute(1, 2, 0))
#     plt.axis("off")
#     break
#
# # plt.show()

###DM
class SimpleDiffusion:
    def __init__(
            self,
            num_diffusion_timesteps=1000,
            img_shape=(3, 64, 64),
            device="cpu",
    ):
        self.num_diffusion_timesteps = num_diffusion_timesteps
        self.img_shape = img_shape
        self.device = device

        self.initialize()

    def initialize(self):
        # BETAs & ALPHAs required at different places in the Algorithm.
        self.beta = self.get_betas()
        self.alpha = 1 - self.beta

        self_sqrt_beta = torch.sqrt(self.beta)
        self.alpha_cumulative = torch.cumprod(self.alpha, dim=0)
        self.sqrt_alpha_cumulative = torch.sqrt(self.alpha_cumulative)
        self.one_by_sqrt_alpha = 1. / torch.sqrt(self.alpha)
        self.sqrt_one_minus_alpha_cumulative = torch.sqrt(1 - self.alpha_cumulative)

    def get_betas(self):
        """linear schedule, proposed in original ddpm paper"""
        scale = 1000 / self.num_diffusion_timesteps
        beta_start = scale * 1e-4
        beta_end = scale * 0.02
        return torch.linspace(
            beta_start,
            beta_end,
            self.num_diffusion_timesteps,
            dtype=torch.float32,
            device=self.device,
        )


def forward_diffusion(sd: SimpleDiffusion, x0: torch.Tensor, timesteps: torch.Tensor):
    """
    前向扩散过程的函数。

    Args:
        sd: SimpleDiffusion 对象,表示简单的扩散过程。
        x0: 原始图像数据。
        timesteps: 一个张量,表示扩散过程中的时间步。

    Returns:
        sample: 扩散后的图像。
        eps: 高斯噪声。
    """
    eps = torch.randn_like(x0)  # 生成与原始图像形状相同的高斯噪声
    mean = get(sd.sqrt_alpha_cumulative, t=timesteps) * x0  # 根据时间步计算扩散过程中的均值
    std_dev = get(sd.sqrt_one_minus_alpha_cumulative, t=timesteps)  # 根据时间步计算扩散过程中的标准差
    sample = mean + std_dev * eps  # 将均值和标准差按比例添加到原始图像上,得到扩散后的图像

    return sample, eps # 返回扩散后的图像和高斯噪声以及经过嵌入后的标签

###可视化前向扩散过程
# # 创建一个 SimpleDiffusion 对象,指定扩散时间步数和设备
# sd = SimpleDiffusion(num_diffusion_timesteps=TrainingConfig.TIMESTEPS, device="cpu")
#
# # 创建一个数据加载器的迭代器,用于加载图像数据
# loader = iter(
#     get_dataloader(
#         dataset_name=BaseConfig.DATASET,
#         batch_size=6,
#         device="cpu",
#     )
# )
#
# x0s, _ = next(loader)  # 从数据加载器中获取一批图像数据
#
# noisy_images = []  # 用于存储扩散后的图像
# specific_timesteps = [0, 10, 50, 100, 150, 200, 250, 300, 400, 600, 800, 999]  # 要可视化的特定时间步
#
# for timestep in specific_timesteps:
#     timestep = torch.as_tensor(timestep, dtype=torch.long)  # 将时间步转换为张量格式
#
#     xts, _ = forward_diffusion(sd, x0s, timestep)  # 对原始图像进行扩散
#     xts = inverse_transform(xts) / 255.0  # 对扩散后的图像进行反向变换和归一化处理
#     xts = make_grid(xts, nrow=1, padding=1)  # 将扩散后的图像拼接成一个网格图像
#
#     noisy_images.append(xts)  # 将拼接后的图像添加到列表中
#
# # 使用 Matplotlib 库绘制扩散过程的可视化结果
# _, ax = plt.subplots(1, len(noisy_images), figsize=(10, 5), facecolor="white")
#
# for i, (timestep, noisy_sample) in enumerate(zip(specific_timesteps, noisy_images)):
#     ax[i].imshow(noisy_sample.squeeze(0).permute(1, 2, 0))  # 在对应的子图中显示扩散后的图像
#     ax[i].set_title(f"t={timestep}", fontsize=8)  # 设置子图的标题为时间步
#     ax[i].axis("off")  # 关闭子图的坐标轴
#     ax[i].grid(False)  # 关闭子图的网格线
#
# plt.suptitle("Forward Diffusion Process", y=0.9)  # 设置整个图形的标题
# plt.axis("off")  # 关闭整个图形的坐标轴
# #plt.show()  # 显示绘制的图形





###训练
@dataclass
class ModelConfig:
    BASE_CH = 64  # 64, 128, 256, 256
    BASE_CH_MULT = (1, 2, 4, 4) # 32, 16, 8, 8
    APPLY_ATTENTION = (False, True, True, False) #(False, True, True, False)
    DROPOUT_RATE = 0.1
    TIME_EMB_MULT = 4 # 128



def train_one_epoch(model, sd, loader, optimizer, scaler, loss_fn, epoch, current_round, user_id,
                    base_config=BaseConfig(), training_config=TrainingConfig()):
    """
    训练一个epoch的函数。

    Args:
        model: 要训练的模型。
        sd: 数据扩散器。
        loader: 数据加载器。
        optimizer: 优化器。
        scaler: 梯度缩放器,用于自动混合精度(AMP)训练。
        loss_fn: 损失函数。
        epoch: 当前的epoch数。
        base_config: 基本配置对象。
        training_config: 训练配置对象。

    Returns:
        mean_loss: 整个epoch的平均损失。
    """
    loss_record = MeanMetric()  # 创建一个用于记录平均损失的对象
    model.train()  # 将模型设置为训练模式

    with tqdm(total=len(loader), dynamic_ncols=True) as tq:  # 创建一个进度条
        tq.set_description(f"Train user{user_id + 1} :: round {current_round}/{dm_num_rounds_fd} :: Epoch: {epoch}/{current_total_epochs}")  # 设置进度条的描述

        for x0s, label in loader:  # 对数据加载器中的每一批数据进行处理
            tq.update(1)  # 更新进度条
            x0s = x0s.to(device)
            label = label.to(device)

            # 随机生成一批时间步
            ts = torch.randint(low=1, high=training_config.TIMESTEPS, size=(x0s.shape[0],), device=base_config.DEVICE)

            xts, gt_noise = forward_diffusion(sd, x0s, ts)  # 对输入数据进行前向扩散

            with amp.autocast():  # 使用自动混合精度(AMP)
                pred_noise = model(xts, label, ts)  # 使用模型预测噪声
                loss = loss_fn(gt_noise, pred_noise)  # 计算损失

            optimizer.zero_grad(set_to_none=True)  # 清零梯度
            scaler.scale(loss).backward()  # 对损失进行缩放,并计算梯度

            # scaler.unscale_(optimizer)  # 如果需要的话,可以对优化器进行梯度裁剪
            # torch.nn.utils.clip_grad_norm_(model.parameters(), 1.0)

            scaler.step(optimizer)  # 更新优化器
            scaler.update()  # 更新scaler

            loss_value = loss.detach().item()  # 获取损失值
            loss_record.update(loss_value)  # 记录损失

            tq.set_postfix_str(s=f"Loss: {loss_value:.4f}")  # 在进度条中显示当前批次的损失

        mean_loss = loss_record.compute().item()  # 计算整个epoch的平均损失

        tq.set_postfix_str(s=f"Epoch Loss: {mean_loss:.4f}")  # 在进度条中显示整个epoch的平均损失

    return mean_loss  # 返回整个epoch的平均损失




@torch.inference_mode()
def reverse_diffusion(model, vae_model, sd, timesteps=1000, img_shape=(3, 64, 64),
                      num_images=10, nrow=10, device="cpu",num_classes=10 ,is_latent = False,  **kwargs):
    """
    反向扩散过程的函数,用于从随机噪声生成图像。

    Args:
        model: 训练好的条件扩散模型。
        sd: SimpleDiffusion 对象,表示简单的扩散过程。
        timesteps: 反向扩散过程的总时间步数。
        img_shape: 生成图像的形状。
        num_images: 每个类别要生成的图像数量。
        nrow: 在生成的网格图像中每行显示的图像数量。
        device: 使用的设备(CPU 或 GPU)。
        **kwargs: 其他可选参数,如 generate_video 和 save_path。

    Returns:
        None
    """
    num_classes = num_classes  # 类别数量
    total_images = num_images * num_classes  # 总图像数量
    x = torch.randn((total_images, *img_shape), device=device)  # 初始化随机噪声

    # 创建类别标签，每个类别生成 num_images 个图像
    y = torch.arange(num_classes, device=device).repeat_interleave(num_images)

    model.eval()  # 将模型设置为评估模式
    vae_model.eval()

    if kwargs.get("generate_video", False):
        outs = []  # 用于存储生成的图像帧

    # 反向迭代时间步
    for time_step in tqdm(iterable=reversed(range(1, timesteps)),
                          total=timesteps - 1, dynamic_ncols=False,
                          desc="Sampling :: ", position=0):
        ts = torch.ones(total_images, dtype=torch.long, device=device) * time_step  # 创建当前时间步的张量
        z = torch.randn_like(x) if time_step > 1 else torch.zeros_like(x)  # 创建随机噪声或零噪声

        predicted_noise = model(x, y, ts)  # 使用模型预测噪声

        # 计算反向扩散过程中的系数
        beta_t = get(sd.beta, ts)
        one_by_sqrt_alpha_t = get(sd.one_by_sqrt_alpha, ts)
        sqrt_one_minus_alpha_cumulative_t = get(sd.sqrt_one_minus_alpha_cumulative, ts)

        # 根据预测的噪声和系数更新图像
        x = (
            one_by_sqrt_alpha_t
            * (x - (beta_t / sqrt_one_minus_alpha_cumulative_t) * predicted_noise)
            + torch.sqrt(beta_t) * z
        )




        if kwargs.get("generate_video", False):
            if is_latent:
                x_decoded = vae_model.decode(x)
                x_inv = inverse_transform(x_decoded).type(torch.uint8)  # 对生成的图像进行反向变换
            else:
                x_inv = inverse_transform(x).type(torch.uint8)
            # x_inv = inverse_transform(x_decoded).type(torch.uint8)  # 对生成的图像进行反向变换
            grid = make_grid(x_inv, nrow=nrow, pad_value=255.0).to("cpu")  # 创建图像网格
            ndarr = torch.permute(grid, (1, 2, 0)).numpy()[:, :, ::-1]  # 调整图像的维度顺序
            outs.append(ndarr)  # 将图像帧添加到列表中

    if kwargs.get("generate_video", False):  # 如果需要生成视频
        frames2vid(outs, kwargs['save_path'])  # 将图像帧转换为视频并保存
        display(Image.fromarray(outs[-1][:, :, ::-1]))  # 显示反向扩散过程最后一步的图像
        return None

    else:  # 如果不生成视频,只保存和显示最后一步的图像
        if is_latent:
            x_decoded = vae_model.decode(x)
            x_inv = inverse_transform(x_decoded).type(torch.uint8)  # 对生成的图像进行反向变换

        else:
            x_inv = inverse_transform(x).type(torch.uint8)
        # x = inverse_transform(x_decoded).type(torch.uint8)  # 对生成的图像进行反向变换
        grid = make_grid(x_inv, nrow=nrow, pad_value=255.0).to("cpu")  # 创建图像网格
        pil_image = TF.functional.to_pil_image(grid)  # 将图像转换为 PIL 图像
        save_path = kwargs['save_path']
        pil_image.save(kwargs['save_path'], format=save_path[-3:].upper())  # 保存图像
        display(pil_image)  # 显示图像
        return None


##K个用户用有标签数据联邦训练分类器，resnet18结构
class ResNetClassifier(nn.Module):
    def __init__(self, num_classes):
        super(ResNetClassifier, self).__init__()
        self.resnet = models.resnet18(pretrained=False)
        # self.resnet = models.resnet18(pretrained=True)
        self.fc = nn.Linear(1000, num_classes)  # 假设输出类别数为num_classes

    def forward(self, x):
        x = self.resnet(x)
        x = F.relu(x)
        x = self.fc(x)
        return x


def train_classifier(model, trainloader, testloader, optimizer, criterion, num_epochs, user_id):
    model.to(device)

    for epoch in range(num_epochs):
        model.train()
        running_loss = 0.0
        correct_train = 0
        total_train = 0
        pbar = tqdm(enumerate(trainloader), total=len(trainloader))
        for batch_idx, (images, labels) in pbar:
            if images.size(0) <= 1:  # 检查每个批次的大小
                continue
            images = images.to(device)
            labels = labels.to(device)

            optimizer.zero_grad()
            outputs = model(images)
            loss = criterion(outputs, labels)
            loss.backward()
            optimizer.step()
            running_loss += loss.item()

            _, predicted = outputs.max(1)
            total_train += labels.size(0)
            correct_train += predicted.eq(labels).sum().item()

            pbar.set_description(
                'User {} Epoch [{}/{}], Loss: {:.4f}'.format(user_id+1, epoch + 1, num_epochs, running_loss / (batch_idx + 1)))

        train_accuracy = 100. * correct_train / total_train
        test_accuracy = test_classifier(model, testloader)

        print(f'  Training Accuracy: {train_accuracy:.2f}%')
        print(f'  Test Accuracy: {test_accuracy:.2f}%')

    print('Training complete.')

    return model.state_dict(), test_accuracy


def test_classifier(model, testloader):
    model.eval()
    correct = 0
    total = 0
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    with torch.no_grad():
        for images, labels in testloader:
            images = images.to(device)
            labels = labels.to(device)

            outputs = model(images)
            _, predicted = torch.max(outputs.data, 1)
            total += labels.size(0)
            correct += (predicted == labels).sum().item()

    accuracy = 100 * correct / total
    # print('Accuracy on test set: {:.2f}%'.format(accuracy))
    return accuracy

def weight_init(m):
    if isinstance(m, nn.Conv2d) or isinstance(m, nn.Linear):
        nn.init.xavier_uniform_(m.weight)
        if m.bias is not None:
            nn.init.constant_(m.bias, 0.0)


def evaluate_model(model, dataloader):
    model.eval()
    correct = 0
    total = 0
    all_predictions = []

    with torch.no_grad():
        for inputs, labels in dataloader:
            inputs = inputs.to(device)
            labels = labels.to(device)
            outputs = model(inputs)
            _, predicted = torch.max(outputs, 1)
            total += labels.size(0)
            correct += (predicted == labels).sum().item()
            all_predictions.extend(predicted.cpu().numpy())

    accuracy = correct / total
    return accuracy, all_predictions


def create_dm_pseudoloader(user_loader):
    for inputs, current_labels in user_loader:
        transformed_inputs = torch.stack([transform_dm(img) for img in inputs])
        yield transformed_inputs, current_labels



def calculate_images_to_generate(train_user_datasets_labelled, train_user_datasets_unlabelled, train_labels, factor):
    images_to_generate = []

    for user_id in range(K):
        # 获取该用户的有标签和无标签数据
        labelled_data = train_user_datasets_labelled[user_id]
        unlabelled_data = train_user_datasets_unlabelled[user_id]

        # 计算该用户的总数据量
        total_data_count = (len(labelled_data) + len(unlabelled_data))
        factor_data_count = round((len(labelled_data) + len(unlabelled_data)) * factor)

        # 计算每个类别应该有的图片数量
        target_count_per_class = factor_data_count // num_classes

        # 统计当前有标签数据中每个类别的数量
        current_class_count = [0] * num_classes
        for data_index in labelled_data:
            class_label = train_labels[data_index]
            current_class_count[class_label] += 1

        # 计算每个类别需要生成的图片数量
        generation_count = [max(0, target_count_per_class - count) for count in current_class_count]

        images_to_generate.append(generation_count)

        # 打印结果
        print(f"User {user_id + 1}:")
        print(f"  Total data count: {total_data_count}")
        print(f"  Factor data count: {factor_data_count}")
        print(f"  Target count per class: {target_count_per_class}")
        print(f"  Current class count: {current_class_count}")
        print(f"  Images to generate per class: {generation_count}")
        print(f"  Total images to generate: {sum(generation_count)}")
        print()

    return images_to_generate


def reverse_diffusion_new(model, vae_model, sd, timesteps=1000, img_shape=(3, 64, 64),
                          num_images=10, device="cpu", num_classes=10, is_latent=False, label=0):
    """
    反向扩散过程的函数,用于从随机噪声生成图像并返回生成的图像。

    Args:
        model: 训练好的条件扩散模型。
        vae_model: VAE模型，用于解码潜在表示（如果is_latent为True）。
        sd: SimpleDiffusion 对象,表示简单的扩散过程。
        timesteps: 反向扩散过程的总时间步数。
        img_shape: 生成图像的形状。
        num_images: 要生成的图像数量。
        device: 使用的设备(CPU 或 GPU)。
        num_classes: 类别数量。
        is_latent: 是否在潜在空间中生成图像。

    Returns:
        生成的图像张量
    """
    x = torch.randn((num_images, *img_shape), device=device)  # 初始化随机噪声

    # 创建类别标签
    # y = torch.randint(0, num_classes, (num_images,), device=device)
    y = torch.tensor([label]*num_images, device=device)

    model.eval()  # 将模型设置为评估模式
    vae_model.eval()

    # 反向迭代时间步
    for time_step in reversed(range(1, timesteps)):
        ts = torch.ones(num_images, dtype=torch.long, device=device) * time_step  # 创建当前时间步的张量
        z = torch.randn_like(x) if time_step > 1 else torch.zeros_like(x)  # 创建随机噪声或零噪声

        predicted_noise = model(x, y, ts)  # 使用模型预测噪声

        # 计算反向扩散过程中的系数
        beta_t = get(sd.beta, ts)
        one_by_sqrt_alpha_t = get(sd.one_by_sqrt_alpha, ts)
        sqrt_one_minus_alpha_cumulative_t = get(sd.sqrt_one_minus_alpha_cumulative, ts)

        # 根据预测的噪声和系数更新图像
        x = (
            one_by_sqrt_alpha_t
            * (x - (beta_t / sqrt_one_minus_alpha_cumulative_t) * predicted_noise)
            + torch.sqrt(beta_t) * z
        )

    if is_latent:
        x = vae_model.decode(x)

    return x



def generate_and_merge_data(train_user_datasets_labelled, trainset, train_labels, images_to_generate, final_dm_model_save_paths, trained_vae_models, save_path):
    new_dataloaders = []

    for user_id, fl_user_data in enumerate(train_user_datasets_labelled):
        with torch.no_grad():
            # 加载模型
            dm_model = UNet(
                input_channels=TrainingConfig.IMG_SHAPE[0],
                output_channels=TrainingConfig.IMG_SHAPE[0],
                base_channels=ModelConfig.BASE_CH,
                base_channels_multiples=ModelConfig.BASE_CH_MULT,
                apply_attention=ModelConfig.APPLY_ATTENTION,
                dropout_rate=ModelConfig.DROPOUT_RATE,
                time_multiple=ModelConfig.TIME_EMB_MULT,
            ).to(BaseConfig.DEVICE)
            dm_model.load_state_dict(torch.load(final_dm_model_save_paths[user_id]))
            dm_model.eval()

            vae_model = trained_vae_models[user_id].to(BaseConfig.DEVICE)
            vae_model.eval()

            # 获取原始有标签数据
            original_data = torch.stack([trainset[idx][0] for idx in fl_user_data]).cpu()
            original_labels = torch.from_numpy(train_labels[fl_user_data])

            # 生成新图片
            generated_images = []
            generated_labels = []
            num_classes = 10  # CIFAR10 有 10 个类别

            sd = SimpleDiffusion(
                num_diffusion_timesteps=TrainingConfig.TIMESTEPS,
                img_shape=TrainingConfig.IMG_SHAPE,
                device=BaseConfig.DEVICE,
            )
            # 在主循环开始前初始化字典来跟踪每个类别是否已保存
            saved_classes = set()
            for class_label, num_images in enumerate(images_to_generate[user_id]):
                # 分批生成图像
                batch_size = 500  # 可以根据GPU内存调整这个值
                for i in range(0, num_images, batch_size):
                    batch = reverse_diffusion_new(
                        model=dm_model,
                        vae_model=vae_model,
                        sd=sd,
                        timesteps=TrainingConfig.TIMESTEPS,
                        img_shape=TrainingConfig.IMG_SHAPE,
                        num_images=min(batch_size, num_images-i),
                        device=BaseConfig.DEVICE,
                        num_classes=num_classes,
                        is_latent=True,
                        label=class_label#指定生成的类别
                    )
                    generated_images.extend([img.cpu() for img in batch])
                    generated_labels.extend([class_label] * len(batch))

                    # 如果这个类别还没有保存图片，保存第一张
                    if class_label not in saved_classes:
                        save_dir = os.path.join(save_path, f"generated_images/user_{user_id}")
                        os.makedirs(save_dir, exist_ok=True)
                        save_image(batch[0].cpu(), f"{save_dir}/class_{class_label}_sample.png")
                        saved_classes.add(class_label)
                    # 定期清理GPU缓存
                    torch.cuda.empty_cache()


            # 将生成的图片转换为张量
            # generated_images = torch.stack(generated_images)
            # generated_labels = torch.tensor(generated_labels)

            if generated_images:
                generated_images = torch.stack(generated_images)
                generated_labels = torch.tensor(generated_labels)
    
                # 将 generated_images 从 [0, 1] 范围调整到 [-1, 1] 范围
                generated_images = generated_images * 2 - 1
    
                all_images = torch.cat([original_data, generated_images])
                all_labels = torch.cat([original_labels, generated_labels])
            else:
                print(f"No images were generated for user {user_id}")
                all_images = original_data
                all_labels = original_labels

            # 创建新的 Dataset 和 DataLoader
            dataset = TensorDataset(all_images, all_labels)
            dataloader = DataLoader(dataset, batch_size=32, shuffle=True)

            new_dataloaders.append(dataloader)

            # # 可选：保存一些生成的图片用于可视化
            # save_dir =os.path.join(save_path, f"generated_images/user_{user_id}")
            # os.makedirs(save_dir, exist_ok=True)
            # for i in range(min(10, len(generated_images))):
            #     save_image((generated_images[i]+1)/2, f"{save_dir}/class_{generated_labels[i]}_sample_{i}.png")

        # 循环结束后，将模型移回CPU并清理GPU内存
        dm_model.cpu()
        vae_model.cpu()
        torch.cuda.empty_cache()

    return new_dataloaders


def convert_dataset_to_tensor(dataset):
    images, labels = [], []
    for img, lbl in dataset:
        if not isinstance(img, torch.Tensor):
            img = torch.tensor(img)
        if not isinstance(lbl, torch.Tensor):
            lbl = torch.tensor(lbl)
        images.append(img)
        labels.append(lbl)
    
    images = torch.stack(images)
    labels = torch.stack(labels).long()
    return TensorDataset(images, labels)

if __name__ == '__main__':
    ###构造dataloader
    # 定义图像预处理操作
    # 数据范围[-1,1]
    transform_classify = TF.Compose(
        [TF.ToTensor(),
         TF.Normalize((0.5, 0.5, 0.5), (0.5, 0.5, 0.5))])
    # 数据范围[0,1]
    transform_vae = TF.Compose(
        [TF.ToTensor()])
     #数据范围[0,1]
    transform_dm = TF.Compose(
        [
            # TF.ToTensor(),
            TF.Resize((32, 32),
                      interpolation=TF.InterpolationMode.BICUBIC,
                      antialias=True),
            TF.RandomHorizontalFlip(),
            # TF.Lambda(lambda t: (t * 2) - 1) # Scale between [-1, 1]
            TF.Lambda(scale_to_range)
        ]
    )

    # 下载并加载CIFAR-10训练集
    trainset = torchvision.datasets.CIFAR10(root='./data', train=True,
                                            download=True, transform=transform_classify)

    trainset_vae = torchvision.datasets.CIFAR10(root='./data', train=True,
                                                download=True, transform=transform_vae)

    trainset_dm = torchvision.datasets.CIFAR10(root='./data', train=True,
                                               download=True, transform=transform_dm)
    # 下载并加载CIFAR-10测试集
    testset = torchvision.datasets.CIFAR10(root='./data', train=False,
                                           download=True, transform=transform_classify)
    testset_vae = torchvision.datasets.CIFAR10(root='./data', train=False,
                                               download=True, transform=transform_vae)

    # 获取训练集中每个类别的下标
    train_labels = np.array(trainset.targets)
    print(train_labels.shape)
    train_indices_by_class = []
    for class_label in range(10):
        indices = np.where(train_labels == class_label)[0]
        train_indices_by_class.append(indices)

    # 获取测试集中每个类别的下标
    test_labels = np.array(testset.targets)
    print(test_labels.shape)
    test_indices_by_class = []
    for class_label in range(10):
        indices = np.where(test_labels == class_label)[0]
        test_indices_by_class.append(indices)

    # 训练集的一部分数据是有标签的
    beta_data = 0.898
    # 划分train_indices_by_class为有标签数据train_indices_by_class_labelled和无标签数据train_indices_by_class_unlabelled
    train_indices_by_class_labelled = [[] for _ in range(10)]
    train_indices_by_class_unlabelled = [[] for _ in range(10)]
    for class_label, class_indices in enumerate(train_indices_by_class):
        # 计算需要使用的数据数量
        total_data_count = len(class_indices)
        subset_data_count = int(total_data_count * beta_data)

        # 随机选择subset_data_count个数据的下标
        subset_indices = np.random.choice(class_indices, size=subset_data_count, replace=False)

        # 将subset_indices划分为train_indices_by_class_labelled和train_indices_by_class_unlabelled
        train_indices_by_class_labelled[class_label].extend(subset_indices)
        train_indices_by_class_unlabelled[class_label].extend(
            [index for index in class_indices if index not in subset_indices])

    # # 打印划分后的数据数量和下标
    # for class_label in range(10):
    #     print(f"Class {class_label} train_indices_by_class_labelled count:", len(train_indices_by_class_labelled[class_label]))
    #     print(f"Class {class_label} train_indices_by_class_labelled indices:", train_indices_by_class_labelled[class_label])
    #     print(f"Class {class_label} train_indices_by_class_unlabelled count:", len(train_indices_by_class_unlabelled[class_label]))
    #     # print(f"Class {class_label} train_indices_by_class_unlabelled indices:", train_indices_by_class_unlabelled[class_label])

    # print(train_indices_by_class)
    # print(test_indices_by_class)

    # 划分有标签数据的训练集.gamma小，non-iid;gamma大，iid
    gamma_labelled = 0.1
    gamma_unlabelled = 0.1
    train_user_datasets_labelled = [[] for _ in range(K)]
    train_user_datasets_unlabelled = [[] for _ in range(K)]
    test_user_datasets = [[] for _ in range(K)]
    test_user_datasets_vae = [[] for _ in range(K)]
    temp_subset_sizes = [[] for _ in range(10)]
    temp_subset_sizes_vae = [[] for _ in range(10)]
    j = -1
    for indices in train_indices_by_class_labelled:
        j = j + 1
        # 计算迪利克雷分布的alpha参数
        alpha = [1] * K  # 这里假设每个类别的数据划分为K个子集
        alpha = np.array(alpha) * gamma_labelled

        # 生成迪利克雷分布样本
        subset_sizes = np.random.dirichlet(alpha, size=1)[0]
        # print(subset_sizes)
        temp_subset_sizes[j] = subset_sizes

        # 随机打乱类别下标
        np.random.shuffle(indices)

        # 将每个类别的数据划分为K个子集，并分给K个用户
        start_index = 0
        for i in range(K):
            end_index = start_index + int(len(indices) * subset_sizes[i])
            train_user_datasets_labelled[i].extend(indices[start_index:end_index])
            start_index = end_index
    # 打印每个用户的数据数量和每个类别的数据下标及索引,有标签数据

    for user_id, user_data in enumerate(train_user_datasets_labelled):
        print(f"User {user_id + 1} labeled train data count:", len(user_data))
        user_class_count = [0] * 10
        user_class_indices = [[] for _ in range(10)]
        for data_index in user_data:
            class_label = train_labels[data_index]
            user_class_count[class_label] += 1
            user_class_indices[class_label].append(data_index)
        print(f"User {user_id + 1} labeled train class count:", user_class_count)
        # for class_label, class_indices in enumerate(user_class_indices):
        #     print(f"User {user_id+1} unlabeled train class {class_label} indices:", class_indices)
        # # # 查看每个用户的训练数据集
        # # # 创建一个总图，并设置子图的行数和列数
        # # # 计算子图网格的行数和列数
        # num_rows = len(user_class_indices)
        # num_cols = max([len(indices) for indices in user_class_indices])
        # # 创建子图网格
        # fig, axes = plt.subplots(num_rows, num_cols, figsize=(10, 10),gridspec_kw={'wspace': 0.1, 'hspace': 0.1})
        # for class_label, class_indices in enumerate(user_class_indices):
        #      # 在对应的子图中显示每个类别的图像
        #      for i, index in enumerate(class_indices):
        #          image = trainset.data[index]
        #          ax = axes[class_label][i]
        #          ax.imshow(image)
        #          ax.axis('off')
        #
        #      # 删除未使用的子图网格
        #      for j in range(len(class_indices), num_cols):
        #          axes[class_label][j].axis('off')
        #      # 设置子图标题
        #      axes[class_label][0].set_title(f"Class {class_label}")
        #  # 显示总图
        # plt.show()

    j = -1
    for indices in train_indices_by_class_unlabelled:
        j = j + 1
        # 计算迪利克雷分布的alpha参数
        alpha = [1] * K  # 这里假设每个类别的数据划分为K个子集
        alpha = np.array(alpha) * gamma_unlabelled

        # 生成迪利克雷分布样本
        subset_sizes = np.random.dirichlet(alpha, size=1)[0]
        temp_subset_sizes_vae[j] = subset_sizes
        # print(subset_sizes)

        # 随机打乱类别下标
        np.random.shuffle(indices)

        # 将每个类别的数据划分为K个子集，并分给K个用户
        start_index = 0
        for i in range(K):
            end_index = start_index + int(len(indices) * subset_sizes[i])
            train_user_datasets_unlabelled[i].extend(indices[start_index:end_index])
            start_index = end_index

    # print("temp_subset_sizes:",temp_subset_sizes)

    # 打印每个用户的数据数量和每个类别的数据下标及索引,无标签数据
    for user_id, user_data in enumerate(train_user_datasets_unlabelled):
        print(f"User {user_id + 1} unlabeled train data count:", len(user_data))
        user_class_count = [0] * 10
        user_class_indices = [[] for _ in range(10)]
        for data_index in user_data:
            class_label = train_labels[data_index]
            user_class_count[class_label] += 1
            user_class_indices[class_label].append(data_index)
        print(f"User {user_id + 1} unlabeled train class count:", user_class_count)
        # for class_label, class_indices in enumerate(user_class_indices):
        #     print(f"User {user_id+1} unlabeled train class {class_label} indices:", class_indices)



    #每个用户需要用diffusion model生成的图片数量
    image_factor = 1.0
    images_to_generate = calculate_images_to_generate(train_user_datasets_labelled, train_user_datasets_unlabelled, train_labels, factor=image_factor)

    ##测试集数据划分,划分方式和每个用户的有/无标签数据相同
    j = -1
    for indices in test_indices_by_class:
        j = j + 1
        # 生成迪利克雷分布样本
        subset_sizes = temp_subset_sizes[j]
        # print("test:", subset_sizes)
        # 随机打乱类别下标
        np.random.shuffle(indices)

        # 将每个类别的数据划分为K个子集，并分给K个用户
        start_index = 0
        for i in range(K):
            end_index = start_index + int(len(indices) * subset_sizes[i])
            test_user_datasets[i].extend(indices[start_index:end_index])
            start_index = end_index

    # 打印每个用户的数据数量和每个类别的数据下标及索引,测试集数据
    for user_id, user_data in enumerate(test_user_datasets):
        print(f"User {user_id + 1} test data count:", len(user_data))
        user_class_count = [0] * 10
        user_class_indices = [[] for _ in range(10)]
        for data_index in user_data:
            class_label = test_labels[data_index]
            user_class_count[class_label] += 1
            user_class_indices[class_label].append(data_index)
        print(f"User {user_id + 1} test class count:", user_class_count)
        # for class_label, class_indices in enumerate(user_class_indices):
        #     print(f"User {user_id+1} unlabeled train class {class_label} indices:", class_indices)

    ##VAE测试集数据划分,划分方式和每个用户的无标签数据+有标签数据相同
    j = -1
    for indices in test_indices_by_class:
        j = j + 1
        # 生成迪利克雷分布样本
        subset_sizes = (temp_subset_sizes_vae[j]+temp_subset_sizes[j])*0.5
        # print("test:", subset_sizes)
        # 随机打乱类别下标
        np.random.shuffle(indices)

        # 将每个类别的数据划分为K个子集，并分给K个用户
        start_index = 0
        for i in range(K):
            end_index = start_index + int(len(indices) * subset_sizes[i])
            test_user_datasets_vae[i].extend(indices[start_index:end_index])
            start_index = end_index

    # 打印每个用户的数据数量和每个类别的数据下标及索引,测试集数据
    for user_id, user_data in enumerate(test_user_datasets_vae):
        print(f"User {user_id + 1} test vae data count:", len(user_data))
        user_class_count = [0] * 10
        user_class_indices = [[] for _ in range(10)]
        for data_index in user_data:
            class_label = test_labels[data_index]
            user_class_count[class_label] += 1
            user_class_indices[class_label].append(data_index)
        print(f"User {user_id + 1} test vae class count:", user_class_count)
        # for class_label, class_indices in enumerate(user_class_indices):
        #     print(f"User {user_id+1} unlabeled train class {class_label} indices:", class_indices)

    # 创建每个用户的trainloader
    fl_trainloaders = []
    for fl_user_data in train_user_datasets_labelled:
        # 创建子集数据集对象
        subset = Subset(trainset, fl_user_data)
        # print("fluserdata:",fl_user_data)
        # 创建数据加载器
        trainloader = DataLoader(subset, batch_size=batch_size, shuffle=True)
        # 添加到trainloaders列表
        fl_trainloaders.append(trainloader)

    # 验证是否为每个用户生成了trainloader
    if len(fl_trainloaders) == K:
        print("FL_Trainloader has been generated for each user.K = ", K)
    else:
        print("FL_Trainloader has not been generated for each user.")

    # # 创建每个用户的pseudoloader(训练分类器时的无标签数据集)
    # fl_pseudoloaders = []
    # for fl_user_data in train_user_datasets_unlabelled:
    #     # 创建子集数据集对象
    #     subset = Subset(trainset_dm, fl_user_data)
    #     # print("fluserdata:",fl_user_data)
    #     # 创建数据加载器
    #     pseudoloader = DataLoader(subset, batch_size=batch_size_pseudo, shuffle=True, pin_memory=True,
    #                               num_workers=2)
    #     # 添加到pseudoloaders列表
    #     fl_pseudoloaders.append(pseudoloader)
    #
    # # 验证是否为每个用户生成了pseudoloader
    # if len(fl_pseudoloaders) == K:
    #     print("FL_pseudoloader has been generated for each user.K = ", K)
    # else:
    #     print("FL_pseudoloader has not been generated for each user.")

    # 创建每个用户的pseudoloader(训练分类器时的无标签数据集)
    fl_pseudoloaders = []
    for fl_user_data in train_user_datasets_unlabelled:
        # 创建子集数据集对象
        subset = Subset(trainset, fl_user_data)
        # print("fluserdata:",fl_user_data)
        # 创建数据加载器
        pseudoloader = DataLoader(subset, batch_size=batch_size_pseudo, shuffle=True, pin_memory=True,
                                  num_workers=2)
        # 添加到pseudoloaders列表
        fl_pseudoloaders.append(pseudoloader)

    # 验证是否为每个用户生成了pseudoloader
    if len(fl_pseudoloaders) == K:
        print("FL_pseudoloader has been generated for each user.K = ", K)
    else:
        print("FL_pseudoloader has not been generated for each user.")

    # 创建每个用户的testloader
    fl_testloaders = []
    for fl_user_data in test_user_datasets:
        # 创建子集数据集对象
        subset = Subset(testset, fl_user_data)
        # 创建数据加载器
        testloader = DataLoader(subset, batch_size=batch_size_test, shuffle=True)
        # 添加到trainloaders列表
        fl_testloaders.append(testloader)

    # 验证是否为每个用户生成了testloader
    if len(fl_testloaders) == K:
        print("FL_Testloader has been generated for each user.K = ", K)
    else:
        print("FL_Testloader has not been generated for each user.")

    #创建整个测试集的dataloader
    full_test_dataloader = DataLoader(testset, batch_size=batch_size_test,shuffle=True)
    print("Total_Testloader has not been generated for each user.")
    ###################################################################
    """
    dataset列表：
    train_user_datasets_labelled = [[] for _ in range(K)] 每个用户的有标签训练数据[-1,1]
    train_user_datasets_unlabelled = [[] for _ in range(K)] 每个用户的无标签训练数据[-1,1]
    test_user_datasets = [[] for _ in range(K)] 每个用户的测试数据（分布和有标签数据相同）,用于监控分类器训练程度
    test_user_datasets_vae = [[] for _ in range(K)] 每个用户的VAE测试数据（因为训练VAE不需要label，因此VAE的测试数据分布和无标签数据+有标签数据相同）

    （1）训练分类器：每个用户用有标签数据训练分类器，在测试数据上监控，每隔一定的训练轮次，在full测试集上测试分类精度
    （2）训练VAE：每个用户用有标签数据和无标签数据训练VAE，在VAE测试数据上监控
    （3）训练condition-DM：每个用户用有标签数据+伪标签数据训练condition-DM




    """
    ###############################################
    # 非联邦本地训练分类器

    # 定义损失函数和优化器
    # classify_criterion = nn.CrossEntropyLoss()
    # print("开始非联邦本地训练分类器")
    #
    # # 初始化本地模型
    # local_models = [ResNetClassifier(num_classes).to(device) for _ in range(K)]
    # local_optimizers = [optim.SGD(model.parameters(), lr=0.0001, momentum=0.9) for model in local_models]
    #
    # # 本地训练
    # local_accuracies_pseudo = []
    # local_accuracies_test = []
    # local_accuracies_full_test = []
    #
    # for user_id in range(K):
    #     print(f"训练用户 {user_id + 1} 的本地模型")
    #     trainloader = fl_trainloaders[user_id]
    #     testloader = fl_testloaders[user_id]
    #
    #     # 训练本地模型
    #     train_classifier(local_models[user_id], trainloader, testloader, local_optimizers[user_id], classify_criterion,
    #                      classify_num_rounds_lc, user_id)
    #
    #     # 评估本地模型在 pseudoloader 上的性能
    #     accuracy_pseudo, _ = evaluate_model(local_models[user_id], fl_pseudoloaders[user_id])
    #     local_accuracies_pseudo.append(accuracy_pseudo)
    #
    #
    #     # 评估本地模型在 testloader 上的性能
    #     accuracy_test, _ = evaluate_model(local_models[user_id], testloader)
    #     local_accuracies_test.append(accuracy_test)
    #
    #
    #     # 评估本地模型在 total_testloader 上的性能
    #     accuracy_full_test, _ = evaluate_model(local_models[user_id], full_test_dataloader)
    #     local_accuracies_full_test.append(accuracy_full_test)
    #
    #
    # print("非联邦本地训练分类器完成。")
    #
    # # 输出本地训练性能
    # print("本地训练性能：")
    # total_local_accuracy_pseudo = sum(local_accuracies_pseudo)
    # total_local_accuracy_test = sum(local_accuracies_test)
    # total_local_accuracy_full_test = sum(local_accuracies_full_test)
    #
    # for user in range(K):
    #     print(f"本地训练：用户{user + 1}的分类精度：")
    #     print(f"  在 pseudoloader 上: {local_accuracies_pseudo[user]:.6f}")
    #     print(f"  在 testloader 上: {local_accuracies_test[user]:.6f}")
    #     print(f"  在 testloader 上: {local_accuracies_full_test[user]:.6f}")
    #
    # print(f"本地训练：所有用户在 pseudoloader 上的平均分类精度为: {total_local_accuracy_pseudo / K:.6f}")
    # print(f"本地训练：所有用户在 testloader 上的平均分类精度为: {total_local_accuracy_test / K:.6f}")
    # print(f"本地训练：所有用户在 total_testloader 上的平均分类精度为: {total_local_accuracy_full_test / K:.6f}")



    ####################################################
    classify_lr = 0.0001

    # 创建Excel工作簿和工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Federation Learning Results"

    # 添加表头
    headers = ["K", "batch_size", "γ1", "γ2", "lr", "β", "fd_epoch", "fd_round", "无标签avg", "测试集avg", "full测试集avg"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')

    # 添加常量值
    ws.cell(row=2, column=1, value=K)
    ws.cell(row=2, column=2, value=batch_size)
    ws.cell(row=2, column=3, value=gamma_labelled)
    ws.cell(row=2, column=4, value=gamma_unlabelled)
    ws.cell(row=2, column=5, value=classify_lr)
    ws.cell(row=2, column=6, value=beta_data)

    row = 2


    # ####################################################
    classify_criterion = nn.CrossEntropyLoss()

    best_accuracy_full_test = 0.
    best_accuracy_avg_pseudo = 0.
    best_model_full_test = None
    best_model_avg_pseudo = None
    best_round_full_test = 0
    best_round_avg_pseudo = 0

    # 用于记录每次评估的精度
    full_test_accuracies = []
    avg_pseudo_accuracies = []
    avg_test_accuracies = []

    # 计算每个用户的数据集大小
    classify_dataset_sizes = [len(loader.dataset) for loader in fl_trainloaders]
    total_size = sum(classify_dataset_sizes)  # 所有用户的数据总量
    # 打印每个用户的数据量
    for user_id, size in enumerate(classify_dataset_sizes):
        print(f"用户 {user_id + 1} 的数据量: {size}")
    # 打印所有用户的数据总量
    print(f"所有用户的数据总量: {total_size}")

    # # K个用户联邦训练分类器
    # # 在联邦学习开始前初始化模型
    # classify_models = [ResNetClassifier(num_classes).to(device) for _ in range(K)]
    # # 定义优化器
    # classify_optimizers = [optim.SGD(model.parameters(), lr=classify_lr, momentum=0.9) for model in classify_models]
    # last_round_test_acc = [0. for _ in range(K)]

    # for classify_round in range(1, classify_num_rounds_fd + 1):
    #     print(f"联邦训练：第{classify_round}轮/{classify_num_rounds_fd}轮")
    #     if classify_round <= classify_num_rounds_1:
    #         classify_current_epochs = classify_num_epochs_1
    #     else:
    #         classify_current_epochs = classify_num_epochs_2

    #     # 每个用户进行本地训练
    #     classify_model_dicts = []
    #     for user_id in range(K):
    #         # 训练本地模型并获取更新后的状态字典
    #         local_model_dict, test_acc = train_classifier(classify_models[user_id], fl_trainloaders[user_id], fl_testloaders[user_id], classify_optimizers[user_id], classify_criterion, classify_current_epochs, user_id)
    #         classify_model_dicts.append(local_model_dict)
    #         print(f'  Last Round Test Accuracy: {last_round_test_acc[user_id]:.2f}%')
    #         last_round_test_acc[user_id] = test_acc

    #     # 进行联邦平均
    #     classify_global_dict = federated_average(classify_model_dicts, classify_dataset_sizes)

    #     # 将全局模型参数分发给所有本地模型
    #     for model in classify_models:
    #         model.load_state_dict(classify_global_dict)

    #     if classify_round % 100 == 0:
    #         # 评估全局模型在 full_testloader 上的性能
    #         global_model = classify_models[0]
    #         accuracy_full_test, _ = evaluate_model(global_model, full_test_dataloader)
    #         full_test_accuracies.append(accuracy_full_test)

    #         fed_accuracies_pseudo = []
    #         fed_accuracies_test = []

    #         for user_id in range(K):
    #             pseudoloader = fl_pseudoloaders[user_id]
    #             testloader = fl_testloaders[user_id]
    #             local_model = classify_models[user_id]

    #             # 评估在 pseudoloader 上的性能
    #             accuracy_pseudo, _ = evaluate_model(local_model, pseudoloader)
    #             fed_accuracies_pseudo.append(accuracy_pseudo)

    #             # 评估在 testloader 上的性能
    #             accuracy_test, _ = evaluate_model(local_model, testloader)
    #             fed_accuracies_test.append(accuracy_test)

    #         avg_accuracy_pseudo = sum(fed_accuracies_pseudo) / K
    #         avg_accuracy_test = sum(fed_accuracies_test) / K
    #         avg_pseudo_accuracies.append(avg_accuracy_pseudo)
    #         avg_test_accuracies.append(avg_accuracy_test)

    #         # 将结果写入Excel
    #         ws.cell(row=row, column=7, value=1)  # fd_epoch 固定为1
    #         ws.cell(row=row, column=8, value=classify_round)
    #         ws.cell(row=row, column=9, value=avg_accuracy_pseudo)
    #         ws.cell(row=row, column=10, value=avg_accuracy_test)
    #         ws.cell(row=row, column=11, value=accuracy_full_test)

    #         row += 1

    #         print(f"Round {classify_round}:")
    #         print(f"  Full Test Accuracy: {accuracy_full_test:.6f}")
    #         print(f"  Avg Pseudo Accuracy: {avg_accuracy_pseudo:.6f}")
    #         print(f"  Avg Test Accuracy: {avg_accuracy_test:.6f}")

    #         # 更新最佳模型
    #         if accuracy_full_test > best_accuracy_full_test:
    #             best_accuracy_full_test = accuracy_full_test
    #             best_model_full_test = copy.deepcopy(global_model.state_dict())
    #             best_round_full_test = classify_round

    #         if avg_accuracy_pseudo > best_accuracy_avg_pseudo:
    #             best_accuracy_avg_pseudo = avg_accuracy_pseudo
    #             best_model_avg_pseudo = copy.deepcopy(global_model.state_dict())
    #             best_round_avg_pseudo = classify_round

    # print("联邦训练完成。输出评估结果：")
    # for i, (full_test_acc, avg_pseudo_acc, avg_test_acc) in enumerate(zip(full_test_accuracies, avg_pseudo_accuracies, avg_test_accuracies)):
    #     print(f"Evaluation {i + 1}:")
    #     print(f"  Full Test Accuracy: {full_test_acc:.6f}")
    #     print(f"  Avg Pseudo Accuracy: {avg_pseudo_acc:.6f}")
    #     print(f"  Avg Test Accuracy: {avg_test_acc:.6f}")

    # print(f"\n最佳Full Test模型 - Round: {best_round_full_test}, Accuracy: {best_accuracy_full_test:.6f}")
    # print(f"最佳Avg Pseudo模型 - Round: {best_round_avg_pseudo}, Accuracy: {best_accuracy_avg_pseudo:.6f}")

    # # 保存最佳模型
    load_classify_path = f"/mnt/wzw/CIFAR10_DIR_DIR_CLASSIFY/10/classify"
    load_vae_path = f"/mnt/wzw/CIFAR10_DIR_DIR_VAE/1/vae_beta=0.898"
    source_path = create_incremental_folder(f"/mnt/wzw/CIFAR10_DIR_DIR_DM")
    # classify_save_path = os.path.join(source_path, f"classify")
    classify_combin_save_path = os.path.join(source_path, f"combin_trained_dm_beta={beta_data}")
    # os.makedirs(classify_save_path, exist_ok=True)
    # os.makedirs(classify_combin_save_path, exist_ok=True)

    # torch.save(best_model_full_test, os.path.join(classify_save_path, f'best_model_full_test_global_round{best_round_full_test}.pth'))
    # torch.save(best_model_avg_pseudo, os.path.join(classify_save_path, f'best_model_avg_pseudo_global_round{best_round_avg_pseudo}.pth'))

    # #保存最新的模型
    # torch.save(classify_models[0].state_dict(),os.path.join(classify_save_path, f'final_model_round{classify_num_rounds_fd}.pth'))


    # # 保存Excel文件
    # excel_path = os.path.join(classify_save_path, f'beta={beta_data}.xlsx')
    # wb.save(excel_path)
    # print(f"结果已保存到: {excel_path}")

    # # 清理不再需要的模型和优化器
    # del classify_models
    # del classify_optimizers
    # torch.cuda.empty_cache()


    # 利用联邦训练的分类器构建伪标签数据集
    # 创建new_pseudoloaders，构成伪标签数据集
    # # 加载最佳模型（使用avg_pseudo模型，因为它在伪标签数据集上表现最好）
    # best_model_path = os.path.join(classify_save_path, f'best_model_avg_pseudo_global_round{best_round_avg_pseudo}.pth')
    # 加载最终模型
    best_model_path = os.path.join(load_classify_path, f'beta=0.898_best_model_avg_pseudo_global_round2675.pth')
    best_model_state_dict = torch.load(best_model_path)

    # 创建一个新的ResNetClassifier实例并加载最佳模型状态
    best_model = ResNetClassifier(num_classes).to(device)
    best_model.load_state_dict(best_model_state_dict)
    best_model.eval()

    # 利用联邦训练的分类器构建伪标签数据集,伪标签数据集的数据范围为[-1,1],为了符合DM的数据输入范围，将数据转换到[0,1]
    new_pseudoloaders = []

    for user in range(K):
        pseudoloader = fl_pseudoloaders[user]
        user_inputs = []
        user_labels = []

        with torch.no_grad():
            for inputs, _ in pseudoloader:
                inputs = inputs.to(device)
                outputs = best_model(inputs)
                _, predicted = torch.max(outputs.data, 1)

                user_inputs.extend(inputs.cpu())
                user_labels.extend(predicted.cpu())

        user_inputs = torch.stack(user_inputs)
        user_labels = torch.tensor(user_labels)

        assert len(user_inputs) == len(user_labels), f"用户 {user} 的输入数据和标签数量不匹配"

        user_dataset = TensorDataset(user_inputs, user_labels)
        user_dataloader = DataLoader(user_dataset, batch_size=batch_size_dm, shuffle=True)
        new_pseudoloaders.append(user_dataloader)

        # 删除不再需要的中间变量
        del user_inputs, user_labels, user_dataset
        torch.cuda.empty_cache()

    print(f"创建了 {len(new_pseudoloaders)} 个新的伪标签 DataLoader")

    # 打印每个用户的新伪标签数据集大小
    for user, loader in enumerate(new_pseudoloaders):
        print(f"用户 {user + 1} 的新伪标签数据集大小: {len(loader.dataset)}")

    # 将有标签数据集和伪标签数据集合并，并将数据范围转换到[0,1]之间
    combination_dmloaders = []

    for fl_loader, pseudo_loader in zip(fl_trainloaders, new_pseudoloaders):
        # 获取两个数据集
        fl_dataset = convert_dataset_to_tensor(fl_loader.dataset)
        pseudo_dataset = convert_dataset_to_tensor(pseudo_loader.dataset)
        
        # 合并数据集
        combined_dataset = ConcatDataset([fl_dataset, pseudo_dataset])
        
        # 创建新的 DataLoader
        new_batch_size = max(fl_loader.batch_size, pseudo_loader.batch_size)  # 或者选择其他合适的 batch size
        combined_loader = DataLoader(combined_dataset, batch_size=new_batch_size, shuffle=True)
        
        combination_dmloaders.append(combined_loader)

    print(f"创建了 {len(combination_dmloaders)} 个合并的 DataLoader")

     # 打印每个用户的合并标签数据集大小
    for user, loader in enumerate(combination_dmloaders):
        print(f"用户 {user + 1} 的合成数据集大小: {len(loader.dataset)}")

    #为diffusion模型创建预处理后的数据加载器,将数据转换到[0,1]之间
    dm_pseudoloaders = [create_dm_pseudoloader(loader) for loader in combination_dmloaders]

    print(f"创建了 {len(dm_pseudoloaders)} 个预处理后的 DataLoader 用于 Diffusion 模型")

    # 清理内存
    gc.collect()
    torch.cuda.empty_cache()

    # # 为diffusion模型创建预处理后的数据加载器
    # dm_pseudoloaders = []
    #
    # for user_loader in new_pseudoloaders:
    #     preprocessed_inputs = []
    #     labels = []
    #
    #     for inputs, current_labels in user_loader:
    #         # 应用转换
    #         transformed_inputs = torch.stack([transform_dm(img) for img in inputs])
    #
    #         preprocessed_inputs.append(transformed_inputs)
    #         labels.append(current_labels)
    #
    #     preprocessed_inputs = torch.cat(preprocessed_inputs)
    #     labels = torch.cat(labels)
    #
    #     diffusion_dataset = TensorDataset(preprocessed_inputs, labels)
    #     diffusion_loader = DataLoader(diffusion_dataset, batch_size=batch_size_dm, shuffle=True)
    #     dm_pseudoloaders.append(diffusion_loader)
    #
    # print(f"创建了 {len(dm_pseudoloaders)} 个预处理后的 DataLoader 用于 Diffusion 模型")
    #
    # # 打印每个用户的预处理后数据集大小
    # for user, loader in enumerate(dm_pseudoloaders):
    #     print(f"用户 {user + 1} 的预处理后数据集大小: {len(loader.dataset)}")




    ###############################################
    # # K个用户联邦训练VAE-GAN
    # vae_model_list = []
    # vae_optimizer_list = []
    # disc_optimizer_list = []
    # for _ in range(K):
    #     vae_model = VAE(latent_c, downsample=downsample_factor).to(device)
    #     vae_model_list.append(vae_model)

    #     vae_params = []
    #     for name, module in vae_model.named_children():
    #         if name != 'discriminator':  # 排除判别器的参数
    #             vae_params.extend(list(module.parameters()))
    #     vae_optimizer = torch.optim.Adam(vae_params, lr=1e-4, eps=1e-08, betas=(beta1, beta2))
    #     # vae_optimizer = torch.optim.Adam(vae_params, lr=2e-4)
    #     vae_optimizer_list.append(vae_optimizer)

    #     disc_optimizer = torch.optim.Adam(vae_model.discriminator.parameters(), lr=1e-4, eps=1e-08,
    #                                       betas=(beta1, beta2))
    #     # disc_optimizer = torch.optim.Adam(vae_model.discriminator.parameters(), lr=2e-4)
    #     disc_optimizer_list.append(disc_optimizer)

    # # vae_model_list_lc = vae_model_list
    # vae_model_list_fd = vae_model_list
    # # source_path = create_incremental_folder(f"FD-VAE-GAN+condition_DM")
    # vae_save_path = os.path.join(source_path, f'vae_beta={beta_data}')
    # # # 本地训练每个用户的VAE
    # # for user_id, user_data in enumerate(train_user_datasets_labelled):
    # #     print(f"Training VAE for User {user_id + 1}")
    # #     labelled_data = user_data
    # #     unlabelled_data = train_user_datasets_unlabelled[user_id]
    # #     combined_data = labelled_data + unlabelled_data
    # #     train_vae(vae_model_list_lc[user_id], user_id, combined_data, vae_num_rounds_lc, vae_optimizer_list[user_id], disc_optimizer_list[user_id], vae_save_path)

    # # 联邦训练VAE
    # for round_index in range(vae_num_rounds_fd):
    #     local_models = []
    #     vae_data_sizes = []  # 用于记录每个用户的数据大小
    #     for user_id, user_data in enumerate(train_user_datasets_labelled):
    #         print(f"Fed Training VAE for User {user_id + 1}")
    #         # model = copy.deepcopy(fed_model[user_id])
    #         labelled_data = user_data
    #         unlabelled_data = train_user_datasets_unlabelled[user_id]
    #         combined_data = labelled_data + unlabelled_data
    #         # 记录当前用户的数据大小
    #         vae_data_sizes.append(len(combined_data))
    #         local_model = train_vae_fed(vae_model_list_fd[user_id], user_id, combined_data, vae_num_epochs, vae_optimizer_list[user_id],
    #                                     disc_optimizer_list[user_id], round_index, vae_save_path)
    #         local_models.append(local_model)

    #         # 清理内存
    #         del labelled_data, unlabelled_data, combined_data
    #         torch.cuda.empty_cache()

    #     # 更新全局模型
    #     vae_global_model = federated_average(local_models, vae_data_sizes)

    #     # 将更新后的全局模型分发给所有用户
    #     for model in vae_model_list_fd:
    #         model.load_state_dict(vae_global_model)

    #     print(f"Completed round {round_index + 1}/{vae_num_rounds_fd}")

    #     # 清理内存
    #     del local_models
    #     torch.cuda.empty_cache()


    # # 测试VAE模型
    # for user_id, user_data in enumerate(test_user_datasets_vae):
    #     vae_test_model = vae_model_list[user_id]
    #     vae_test_loader = DataLoader(Subset(testset_vae, user_data), batch_size=64, shuffle=True)
    #     vae_test_model.eval()
    #     with torch.no_grad():
    #         original_images = []
    #         reconstructed_images = []
    #         for batch_idx, (data, _) in enumerate(vae_test_loader):
    #             data = data.to(device, dtype=torch.float)  # 将输入数据转换为与模型权重相同的类型和设备
    #             recon_batch, _, _ = vae_test_model(data)
    #             original_images.append(data.cpu().numpy())
    #             reconstructed_images.append(recon_batch.cpu().numpy())

    #             # 每处理完一个batch就清理一次缓存
    #             torch.cuda.empty_cache()

    #         original_images = np.concatenate(original_images, axis=0)
    #         reconstructed_images = np.concatenate(reconstructed_images, axis=0)

    #     # 显示前10个原始图像和重构图像
    #     show_images(original_images, reconstructed_images, num_images=10)
    #     # 清理内存
    #     del original_images, reconstructed_images, vae_test_model, vae_test_loader
    #     torch.cuda.empty_cache()

    # # 最终清理
    # del vae_model_list, vae_optimizer_list, disc_optimizer_list
    # torch.cuda.empty_cache()




    ###############################################
    ##训练每个用户的dm
    is_dm_train = True
    is_dm_inference = True
    encoded_dataloaders = []
    tmp_dataloaders = dm_pseudoloaders
    # 假设 `model` 是训练好的VAE模型实例，`dataloader` 是原始的图像dataloader
    # 假设设备为CUDA，如果使用CPU则可以改为 'cpu'
    # fed_vae_model_paths = []
    # for i in range(K):
    #     user_dir = os.path.join(vae_save_path, f'Fed/vae_model/user_{i + 1}')
    #     latest_model = get_latest_model(user_dir, i + 1)
    #     if latest_model:
    #         fed_vae_model_paths.append(latest_model)
    #     else:
    #         print(f"Warning: No final model found for user {i + 1}")
    with torch.no_grad():
        fed_vae_model_paths = []
        for i in range(K):
            user_dir = os.path.join(load_vae_path, f'Fed/vae_model/user_{i + 1}')
            latest_model = get_latest_model(user_dir, i + 1)
            if latest_model:
                fed_vae_model_paths.append(latest_model)
            else:
                print(f"Warning: No final model found for user {i + 1}")

        trained_vae_models = []
        for user_id in range(K):
            vae_model_path = fed_vae_model_paths[user_id]
            vae_model = load_model(vae_model_path, VAE(latent_c, downsample=downsample_factor).to(device))
            vae_model.to(device)
            trained_vae_models.append(vae_model)

    dm_save_path = os.path.join(source_path, f"dm_beta={beta_data}")
    os.makedirs(dm_save_path, exist_ok=True)

    with torch.no_grad():
        for user_id, (user_data, model_path) in enumerate(zip(test_user_datasets_vae, fed_vae_model_paths)):
            trained_vae_model = load_model(model_path, VAE(latent_c, downsample=downsample_factor).to(device))
            trained_vae_model.eval()  # 将模型设置为评估模式
            trained_vae_model.to(device)
            # 创建编码后的dataloader
            encode_dataloader = create_encoded_dataloader_with_labels(trained_vae_model, tmp_dataloaders[user_id], device)
            encoded_dataloaders.append(encode_dataloader)

    # 联邦训练cDM
    if is_dm_train:
        dm_model_list = []
        dm_optimizer_list = []
        for _ in range(K):
            dm_model = UNet(  ##网络采用unet结构
                input_channels=TrainingConfig.IMG_SHAPE[0],  # 输入图像的通道数
                output_channels=TrainingConfig.IMG_SHAPE[0],  # 输出图像的通道数
                base_channels=ModelConfig.BASE_CH,
                base_channels_multiples=ModelConfig.BASE_CH_MULT,
                apply_attention=ModelConfig.APPLY_ATTENTION,
                dropout_rate=ModelConfig.DROPOUT_RATE,
                time_multiple=ModelConfig.TIME_EMB_MULT,
            )
            dm_model.to(BaseConfig.DEVICE)

            dm_optimizer = torch.optim.AdamW(dm_model.parameters(), lr=TrainingConfig.LR)
            dm_model_list.append(dm_model)
            dm_optimizer_list.append(dm_optimizer)

            # dataloader = get_dataloader(
            #     dataset_name=BaseConfig.DATASET,
            #     batch_size=TrainingConfig.BATCH_SIZE,
            #     device=BaseConfig.DEVICE,
            #     pin_memory=True,
            #     num_workers=TrainingConfig.NUM_WORKERS,
            # )

        loss_fn = nn.MSELoss()

        sd = SimpleDiffusion(
            num_diffusion_timesteps=TrainingConfig.TIMESTEPS,
            img_shape=TrainingConfig.IMG_SHAPE,
            device=BaseConfig.DEVICE,
        )

        scaler = amp.GradScaler()

        generate_video = True
        ext = ".mp4" if generate_video else ".png"

        for round_index in range(1, dm_num_rounds_fd + 1):
            dm_local_models = []
            dm_data_sizes = []  # 用于记录每个用户的数据大小
            # 根据当前round序号决定本地训练的epoch数
            if round_index <= dm_num_rounds_1:
                current_total_epochs = dm_num_epochs_1
            else:
                current_total_epochs = dm_num_epochs_2
            for user_id, dm_dataloader in enumerate(encoded_dataloaders):
                # dm_dataloader = encoded_dataloaders[0]
                # vae_model_path = f'vae_model/vae_model_{0}.pt'
                dm_data_sizes.append(len(dm_dataloader))
                # vae_model_path = fed_vae_model_paths[user_id]
                # vae_model = load_model(vae_model_path, VAE(latent_c,downsample=downsample_factor).to(device))
                # vae_model.to(device)
                torch.cuda.empty_cache()  # 清空 CUDA 缓存
                gc.collect()  # 触发 Python 的垃圾回收机制


                for epoch in range(1, current_total_epochs + 1):


                    # 算法 1: 训练
                    train_one_epoch(dm_model_list[user_id], sd, dm_dataloader, dm_optimizer_list[user_id], scaler, loss_fn,
                                    epoch, round_index, user_id)  # 调用 train_one_epoch 函数进行一个 epoch 的训练
                    if (round_index % 50 ==0) or (round_index==1):
                        dm_round_save_path = os.path.join(dm_save_path, f"user{user_id + 1}/round{round_index}")
                        os.makedirs(dm_round_save_path, exist_ok=True)
                        if (epoch == current_total_epochs) or (epoch == 1):  
                            # dm_round_save_path = os.path.join(dm_save_path, f"user{user_id + 1}/round{round_index}")
                            # os.makedirs(dm_round_save_path, exist_ok=True)
                            # log_dir, checkpoint_dir = setup_log_directory(config=BaseConfig())  # 保存路径
                            log_dir = os.path.join(dm_round_save_path,f"Inference" )
                            # checkpoint_dir = os.path.join(dm_round_save_path, f"Checkpoint")
                            os.makedirs(log_dir, exist_ok=True)
                            # os.makedirs(checkpoint_dir, exist_ok=True)

                            dm_pic_save_path = os.path.join(log_dir, f"{epoch}{ext}")  # 构建保存生成图像的路径

                            # 算法 2: 采样
                            reverse_diffusion(dm_model_list[user_id], trained_vae_models[user_id], sd, timesteps=TrainingConfig.TIMESTEPS, num_images=5,
                                              generate_video=generate_video,
                                              save_path=dm_pic_save_path, img_shape=TrainingConfig.IMG_SHAPE, device=BaseConfig.DEVICE,num_classes=num_classes,is_latent=True
                                              )  # 调用 reverse_diffusion 函数进行采样,生成图像


                            # checkpoint_dict = {
                            #     "opt": dm_optimizer_list[user_id].state_dict(),
                            #     "scaler": scaler.state_dict(),
                            #     "model": dm_model_list[user_id].state_dict()
                            # }  # 创建一个字典,包含优化器、scaler 和模型的状态字典
                            # torch.save(checkpoint_dict, os.path.join(checkpoint_dir, "ckpt.tar"))  # 将字典保存到指定路径的文件中
                            # del checkpoint_dict  # 删除字典以释放内存
                    if epoch == current_total_epochs:
                        dm_local_model = dm_model_list[user_id].state_dict()
                        dm_local_models.append(dm_local_model)
            # 更新全局模型
            dm_global_model = federated_average(dm_local_models, dm_data_sizes)
            # 将更新后的全局模型分发给所有用户
            for model in dm_model_list:
                model.load_state_dict(dm_global_model)

            print(f"Completed round {round_index}/{dm_num_rounds_fd}")
        # 在所有联邦训练轮次结束后,保存每个用户的最终模型
        final_dm_model_save_paths = []
        for user_id, dm_model in enumerate(dm_model_list):
            final_model_save_path = os.path.join(dm_save_path, f"final_models/user{user_id + 1}_final_model.pt")
            final_dm_model_save_paths.append(final_model_save_path)

            # 确保保存路径存在
            os.makedirs(os.path.dirname(final_model_save_path), exist_ok=True)

            # 保存模型
            torch.save(dm_model.state_dict(), final_model_save_path)
            print(f"Saved final model for user {user_id + 1} at {final_model_save_path}")

        print("All user models have been saved successfully.")

    if is_dm_inference:
        with torch.no_grad():
            inference_dm_model_list = []
            model = UNet(
                input_channels=TrainingConfig.IMG_SHAPE[0],
                output_channels=TrainingConfig.IMG_SHAPE[0],
                base_channels=ModelConfig.BASE_CH,
                base_channels_multiples=ModelConfig.BASE_CH_MULT,
                apply_attention=ModelConfig.APPLY_ATTENTION,
                dropout_rate=ModelConfig.DROPOUT_RATE,
                time_multiple=ModelConfig.TIME_EMB_MULT,
            )
            sd = SimpleDiffusion(
                num_diffusion_timesteps=TrainingConfig.TIMESTEPS,
                img_shape=TrainingConfig.IMG_SHAPE,
                device=BaseConfig.DEVICE,
            )
            for user_id in range(K):

                # checkpoint_dir=r"D:\Lib\wzw\myProject\FLDM\Logs_Checkpoints\checkpoints\version_14"
                # model.load_state_dict(torch.load(os.path.join(checkpoint_dir, "ckpt.tar"), map_location='cpu')['model'])
                model.load_state_dict(torch.load(final_dm_model_save_paths[user_id]))

                model.to(BaseConfig.DEVICE)



                # log_dir = "inference_results"
                log_dir = os.path.join(dm_save_path, f"inference_results")
                os.makedirs(log_dir, exist_ok=True)
                ##################################
                generate_video = False

                ext = ".mp4" if generate_video else ".png"
                filename = f"user{user_id + 1}_{datetime.now().strftime('%Y%m%d-%H%M%S')}{ext}"

                dm_result_save_path = os.path.join(log_dir, filename)

                reverse_diffusion(
                    model,
                    trained_vae_models[user_id],
                    sd,
                    num_images=10,
                    generate_video=generate_video,
                    save_path=dm_result_save_path,
                    timesteps=1000,
                    img_shape=TrainingConfig.IMG_SHAPE,
                    device=BaseConfig.DEVICE,
                    nrow=10,
                    num_classes=num_classes,
                    is_latent=True
                )
                print(dm_result_save_path)


            del inference_dm_model_list, dm_model_list, model, sd
            torch.cuda.empty_cache()
            gc.collect()


    # ##生成新的图片
    # new_dataloaders = generate_and_merge_data(
    #     train_user_datasets_labelled,
    #     trainset,
    #     train_labels,
    #     images_to_generate,
    #     final_dm_model_save_paths,
    #     trained_vae_models,
    #     classify_combin_save_path,#用于区分combination_trained_dm的结果
    # )

    # #用新的合成数据集进行联邦训练分类器
    # # classify_criterion = nn.CrossEntropyLoss()

    # best_accuracy_full_test = 0.
    # best_accuracy_avg_pseudo = 0.
    # best_model_full_test = None
    # best_model_avg_pseudo = None
    # best_round_full_test = 0
    # best_round_avg_pseudo = 0

    # # 用于记录每次评估的精度
    # full_test_accuracies = []
    # avg_pseudo_accuracies = []
    # avg_test_accuracies = []

    # # 计算每个用户的数据集大小
    # classify_dataset_sizes = [len(loader.dataset) for loader in new_dataloaders]
    # total_size = sum(classify_dataset_sizes)  # 所有用户的数据总量
    # # 打印每个用户的数据量
    # for user_id, size in enumerate(classify_dataset_sizes):
    #     print(f"用户 {user_id + 1} 的合成数据量: {size}")
    # # 打印所有用户的数据总量
    # print(f"所有用户的合成数据总量: {total_size}")

    # # K个用户联邦训练分类器
    # # 在联邦学习开始前初始化模型
    # classify_models = [ResNetClassifier(num_classes).to(device) for _ in range(K)]
    # # 定义优化器
    # classify_optimizers = [optim.SGD(model.parameters(), lr=classify_lr, weight_decay=1e-4, momentum=0.9) for model in classify_models]
    # last_round_test_acc = [0. for _ in range(K)]

    # for classify_round in range(1, classify_num_rounds_fd2 + 1):
    #     print(f"联邦训练：第{classify_round}轮/{classify_num_rounds_fd}轮")
    #     if classify_round <= classify_num_rounds_1:
    #         classify_current_epochs = classify_num_epochs_1
    #     else:
    #         classify_current_epochs = classify_num_epochs_2

    #     # 每个用户进行本地训练
    #     classify_model_dicts = []
    #     for user_id in range(K):
    #         # 训练本地模型并获取更新后的状态字典
    #         local_model_dict, test_acc = train_classifier(classify_models[user_id], new_dataloaders[user_id], fl_testloaders[user_id], classify_optimizers[user_id], classify_criterion, classify_current_epochs, user_id)
    #         classify_model_dicts.append(local_model_dict)
    #         print(f'  Last Round Test Accuracy: {last_round_test_acc[user_id]:.2f}%')
    #         last_round_test_acc[user_id] = test_acc

    #     # 进行联邦平均
    #     classify_global_dict = federated_average(classify_model_dicts, classify_dataset_sizes)

    #     # 将全局模型参数分发给所有本地模型
    #     for model in classify_models:
    #         model.load_state_dict(classify_global_dict)

    #     if classify_round % 25 == 0:
    #         # 评估全局模型在 full_testloader 上的性能
    #         global_model = classify_models[0]
    #         accuracy_full_test, _ = evaluate_model(global_model, full_test_dataloader)
    #         full_test_accuracies.append(accuracy_full_test)

    #         fed_accuracies_pseudo = []
    #         fed_accuracies_test = []

    #         for user_id in range(K):
    #             pseudoloader = fl_pseudoloaders[user_id]
    #             testloader = fl_testloaders[user_id]
    #             local_model = classify_models[user_id]

    #             # 评估在 pseudoloader 上的性能
    #             accuracy_pseudo, _ = evaluate_model(local_model, pseudoloader)
    #             fed_accuracies_pseudo.append(accuracy_pseudo)

    #             # 评估在 testloader 上的性能
    #             accuracy_test, _ = evaluate_model(local_model, testloader)
    #             fed_accuracies_test.append(accuracy_test)

    #         avg_accuracy_pseudo = sum(fed_accuracies_pseudo) / K
    #         avg_accuracy_test = sum(fed_accuracies_test) / K
    #         avg_pseudo_accuracies.append(avg_accuracy_pseudo)
    #         avg_test_accuracies.append(avg_accuracy_test)

    #         # 将结果写入Excel
    #         ws.cell(row=row, column=7, value=1)  # fd_epoch 固定为1
    #         ws.cell(row=row, column=8, value=classify_round)
    #         ws.cell(row=row, column=9, value=avg_accuracy_pseudo)
    #         ws.cell(row=row, column=10, value=avg_accuracy_test)
    #         ws.cell(row=row, column=11, value=accuracy_full_test)

    #         row += 1

    #         print(f"Round {classify_round}:")
    #         print(f"  Full Test Accuracy: {accuracy_full_test:.6f}")
    #         print(f"  Avg Pseudo Accuracy: {avg_accuracy_pseudo:.6f}")
    #         print(f"  Avg Test Accuracy: {avg_accuracy_test:.6f}")

    #         # 更新最佳模型
    #         if accuracy_full_test > best_accuracy_full_test:
    #             best_accuracy_full_test = accuracy_full_test
    #             best_model_full_test = copy.deepcopy(global_model.state_dict())
    #             best_round_full_test = classify_round

    #         if avg_accuracy_pseudo > best_accuracy_avg_pseudo:
    #             best_accuracy_avg_pseudo = avg_accuracy_pseudo
    #             best_model_avg_pseudo = copy.deepcopy(global_model.state_dict())
    #             best_round_avg_pseudo = classify_round

    # print("联邦训练完成。输出评估结果：")
    # for i, (full_test_acc, avg_pseudo_acc, avg_test_acc) in enumerate(zip(full_test_accuracies, avg_pseudo_accuracies, avg_test_accuracies)):
    #     print(f"Evaluation {i + 1}:")
    #     print(f"  Full Test Accuracy: {full_test_acc:.6f}")
    #     print(f"  Avg Pseudo Accuracy: {avg_pseudo_acc:.6f}")
    #     print(f"  Avg Test Accuracy: {avg_test_acc:.6f}")

    # print(f"\n合成数据集：最佳Full Test模型 - Round: {best_round_full_test}, Accuracy: {best_accuracy_full_test:.6f}")
    # print(f"合成数据集：最佳Avg Pseudo模型 - Round: {best_round_avg_pseudo}, Accuracy: {best_accuracy_avg_pseudo:.6f}")

    # # 保存最佳模型
    # # source_path = create_incremental_folder(f"FD-VAE-GAN+condition_DM+classify")
    # # classify_save_path = os.path.join(source_path, f"classify")
    # # os.makedirs(classify_save_path, exist_ok=True)

    # torch.save(best_model_full_test, os.path.join(classify_combin_save_path, f'factor={image_factor}_syn_best_model_full_test_global_round{best_round_full_test}.pth'))
    # torch.save(best_model_avg_pseudo, os.path.join(classify_combin_save_path, f'factor={image_factor}_syn_best_model_avg_pseudo_global_round{best_round_avg_pseudo}.pth'))

    # #保存最新的模型
    # torch.save(classify_models[0].state_dict(),os.path.join(classify_combin_save_path, f'factor={image_factor}_syn_final_model_round{classify_num_rounds_fd}.pth'))


    # # 保存Excel文件
    # excel_path = os.path.join(classify_combin_save_path, f'factor={image_factor}.xlsx')
    # wb.save(excel_path)
    # print(f"结果已保存到: {excel_path}")

